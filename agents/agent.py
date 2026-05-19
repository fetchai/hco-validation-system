import os
import base64
import time
import csv
import logging
import threading
from contextlib import contextmanager
from datetime import datetime, timezone
from uuid import uuid4
from typing import Any, Dict, List, Optional
from io import BytesIO
from uagents import Agent, Context, Protocol, Model
from uagents_core.storage import ExternalStorage
from pydantic import Field
from uagents_core.contrib.protocols.chat import (
    chat_protocol_spec,
    ChatMessage,
    ChatAcknowledgement,
    TextContent,
    ResourceContent,
    StartSessionContent,
    MetadataContent,
)

from image_analysis import get_image_analysis, extract_certificate_data, validate_certificate_in_sheets
from html_certificate_generator import generate_certificate_with_html_templates, generate_export_certificate, generate_slaughterhouse_certificate
from database import init_database, save_certificate_to_db, get_certificate_from_db, get_certificate_file_from_db
from openai import OpenAI
import re
import requests
from bs4 import BeautifulSoup
import pandas as pd
from io import StringIO
import json

import hco_logger as hlog

hlog.configure()
perf_logger = logging.getLogger("hco.perf")
perf_logger.setLevel(logging.INFO)


@contextmanager
def step_timer(step_name: str, timings: Dict[str, float]):
    """Record wall-time of a code block into *timings[step_name]* (seconds)."""
    t0 = time.monotonic()
    try:
        yield
    finally:
        elapsed = time.monotonic() - t0
        timings[step_name] = elapsed
        perf_logger.info("STEP_TIMER %s=%.2fs", step_name, elapsed)


# ---------------------------------------------------------------------------
# In-process async job store for long-running certificate generation
# ---------------------------------------------------------------------------
_generation_jobs: Dict[str, Dict[str, Any]] = {}
_jobs_lock = threading.Lock()

GENERATION_JOB_TTL_S = 3600  # keep completed jobs for 1 hour


def _set_job(job_id: str, data: Dict[str, Any]) -> None:
    with _jobs_lock:
        _generation_jobs[job_id] = data


def _get_job(job_id: str) -> Optional[Dict[str, Any]]:
    with _jobs_lock:
        return _generation_jobs.get(job_id)


def _gc_jobs() -> None:
    """Remove completed/failed jobs older than TTL."""
    now = time.time()
    with _jobs_lock:
        expired = [
            jid for jid, jdata in _generation_jobs.items()
            if jdata.get("status") in ("done", "failed")
            and now - jdata.get("updated_at", 0) > GENERATION_JOB_TTL_S
        ]
        for jid in expired:
            del _generation_jobs[jid]


def _openai_chat_completion_with_retry(
    *,
    model: str,
    messages: List[Dict[str, str]],
    temperature: float,
    max_tokens: int,
    retries: int = 2,
    base_sleep_s: float = 0.8,
):
    """Small retry wrapper to reduce transient OpenAI failures."""
    last_err: Optional[Exception] = None
    for attempt in range(retries + 1):
        try:
            return openai_client.chat.completions.create(
                model=model,
                messages=messages,
                temperature=temperature,
                max_completion_tokens=max_tokens,
            )
        except Exception as e:
            last_err = e
            if attempt < retries:
                time.sleep(base_sleep_s * (2 ** attempt))
                continue
            raise

agent = Agent(
    name="HCO Certificate Validator Agent", 
    seed="HCO Certificate Validator Agent", 
    port=8096, 
    mailbox=True
)

class ImageRequest(Model):
    image_data: str = None  # base64 encoded image (optional)
    filename: str = None    # filename (optional)
    content_type: str = None  # content type (optional)
    text_query: str = None  # text query for certificate verification (optional)

class ImageResponse(Model):
    timestamp: int
    message: str
    agent_address: str
    image_size: int
    processed: bool
    file_data: Optional[str] = None  # base64 encoded file data for downloads
    filename: Optional[str] = None  # filename for downloads
    download_url: Optional[str] = None  # download URL for downloads

class CertificateRequest(Model):
    certificate_no: str
    company_name: str = ""
    company_reg_no: str = ""
    issue_date: str
    certificate_type: str = "halal_certificate"
    certificate_category: str = "halal_certificate"  # halal_certificate, export_meat, export_non_meat, slaughterhouse, domestic
    standards: str = ""
    csv_files_count: int = 0
    text_color: str = "black"
    company_address: str = ""
    pu_au_text: str = ""
    sow: str = "Food Processing and Manufacturing"
    pu: str = ""
    au: str = ""
    pl: str = ""  # Product List content (used for slaughterhouse certificates)
    cert_num_footer: str = ""  # Custom footer text (shows in footer on all pages)
    validity_period: str = "1"
    domestic_logo_1: str = "gac"
    domestic_logo_2: str = "none"
    annex_layout_options: Dict[str, Any] = Field(default_factory=dict)
    # Optional: Microsoft access token (delegated) from a logged-in user.
    # Used to restrict certificate generation to users who can access the OneDrive folder.
    auth_token: str = ""
    # File data fields
    csv_file_data: str = ""  # base64 encoded CSV data (deprecated)
    csv_file_filename: str = ""  # filename (deprecated)
    xlsx_files: List[Dict[str, Any]] = Field(default_factory=list)  # List of {filename: str, data: str} for multiple XLSX files
    # Optional company logo (base64). Expected shape: {filename: str, data: str, content_type: str}
    company_logo: Dict = None

    # Export Certificate fields (common)
    country_of_origin: str = ""
    destination: str = ""
    exporter_name: str = ""
    importer_name: str = ""

    # Export Certificate options
    export_logo_option: str = ""
    export_signature_option: str = ""

    # Meat Export specific fields
    slaughter_date: str = ""
    producer_name: str = ""
    expiry_date: str = ""
    abattoir_address: str = ""
    gross_weight: str = ""
    number_of_carcasses: str = ""
    net_weight: str = ""
    number_of_boxes: str = ""
    batch_reference: str = ""
    halal_cert_number: str = ""
    vet_cert_number: str = ""
    destination_port: str = ""
    loading_port: str = ""
    flight_number: str = ""
    meat_type: str = ""
    awb_number: str = ""
    meat_condition: str = ""
    inspector_name: str = ""

    # Non-Meat Export specific fields
    shipment_mode: str = ""
    invoice_no: str = ""
    vet_health_cert_no: str = ""
    products: List[Dict[str, Any]] = Field(default_factory=list)  # List of product rows for non-meat export
    export_products_per_page: int = 10

class CertificateResponse(Model):
    timestamp: int
    message: str
    agent_address: str
    certificate_id: str
    png_filename: str
    pdf_filename: str
    download_url: str = ""
    csv_logged: bool
    processed: bool
    job_id: str = ""
    async_mode: bool = False


class GenerationStatusRequest(Model):
    job_id: str


class GenerationStatusResponse(Model):
    timestamp: int
    job_id: str
    status: str  # queued | running | done | failed
    message: str = ""
    certificate_id: str = ""
    download_url: str = ""
    timings: Dict[str, float] = Field(default_factory=dict)


class ProductVerifyRequest(Model):
    certificate_no: str
    product_names: List[str] = []
    product_codes: List[str] = []

class ProductVerifyResponse(Model):
    timestamp: int
    message: str
    agent_address: str
    certificate_no: str
    certificate_found: bool = False
    verified: bool = False
    download_url: str = ""
    verified_product_names: List[str] = []
    verified_product_codes: List[str] = []
    missing_product_names: List[str] = []
    missing_product_codes: List[str] = []

class CertificateVerifyRequest(Model):
    certificate_no: str

class CertificateVerifyResponse(Model):
    timestamp: int
    message: str
    agent_address: str
    certificate_no: str
    is_valid: bool
    certificate_data: Dict[str, Any] = Field(default_factory=dict)

class ChatRequest(Model):
    query: str

class ChatResponse(Model):
    timestamp: int
    message: str
    agent_address: str
    query_type: str  # "inquiry" or "verification"
    processed: bool
    certificate_no: Optional[str] = None
    certificate_found: bool = False
    verified_product_names: List[str] = []
    verified_product_codes: List[str] = []
    missing_product_names: List[str] = []
    missing_product_codes: List[str] = []

class CertificateDownloadRequest(Model):
    certificate_no: str
    file_type: str  # "png" or "pdf"

class CertificateDownloadResponse(Model):
    certificate_no: str
    file_data: Optional[str] = None  # base64 encoded file data
    filename: Optional[str] = None  # filename for download
    download_url: Optional[str] = None  # optional OneDrive/SharePoint web URL for large files
    found: bool = False  # Whether certificate file was found
    message: str = ""


class ParseNonMeatDocxRequest(Model):
    filename: str = ""
    file_data: str
    auth_token: str = ""


class ParseNonMeatDocxResponse(Model):
    timestamp: int
    message: str
    agent_address: str
    products: List[Dict[str, Any]] = Field(default_factory=list)
    processed: bool


class AuthValidateRequest(Model):
    access_token: str


class AuthValidateResponse(Model):
    timestamp: int
    message: str
    agent_address: str
    allowed: bool
    user_email: str = ""


class AuthStatusRequest(Model):
    # No fields required; kept as a POST for consistent uAgents REST usage.
    pass


class AuthStatusResponse(Model):
    timestamp: int
    agent_address: str
    authenticated: bool
    expires_at: str = ""


class AuthSyncRequest(Model):
    access_token: str
    expires_at: str = ""
    client_id: str = ""
    user_email: str = ""
    user_name: str = ""


class LoginRequest(Model):
    username: str
    password: str


class LoginResponse(Model):
    timestamp: int
    message: str
    agent_address: str
    token: str = ""
    expires_at: int = 0
    user_email: str = ""
    allowed: bool = False


class AuthSyncResponse(Model):
    timestamp: int
    agent_address: str
    ok: bool
    message: str


class AuthClearRequest(Model):
    # No fields required
    pass


class AuthClearResponse(Model):
    timestamp: int
    agent_address: str
    ok: bool
    message: str


def _parse_csv_env_list(name: str) -> list[str]:
    val = (os.getenv(name) or "").strip()
    if not val:
        return []
    return [v.strip().lower() for v in val.split(",") if v.strip()]


def _graph_me_email(access_token: str) -> str:
    """
    Return a stable email/UPN for the token owner.

    Prefer decoding the JWT locally (does not require Graph User.Read).
    If that fails, fall back to Graph /me (requires User.Read).
    """
    # 1) Try decode JWT claims locally (safe enough because we also use the token to call Graph later)
    try:
        parts = (access_token or "").split(".")
        if len(parts) >= 2:
            import base64 as _b64
            payload_b64 = parts[1]
            payload_b64 += "=" * (-len(payload_b64) % 4)
            payload_json = _b64.urlsafe_b64decode(payload_b64.encode("utf-8")).decode("utf-8")
            claims = json.loads(payload_json) if payload_json else {}
            if isinstance(claims, dict):
                email = (
                    claims.get("preferred_username")
                    or claims.get("upn")
                    or claims.get("email")
                    or claims.get("unique_name")
                    or ""
                )
                if isinstance(email, str) and email.strip():
                    return email.strip()
    except Exception:
        pass

    # 2) Fallback: Graph /me
    resp = requests.get(
        "https://graph.microsoft.com/v1.0/me?$select=mail,userPrincipalName",
        headers={"Authorization": f"Bearer {access_token}"},
        timeout=20,
    )
    resp.raise_for_status()
    data = resp.json() or {}
    return (data.get("mail") or data.get("userPrincipalName") or "").strip()


def _check_user_can_access_onedrive_folder(access_token: str) -> None:
    """
    Raises if the user cannot access the configured OneDrive folder share URL.
    """
    folder_share_url = (
        os.getenv("HCO_ONEDRIVE_FOLDER_SHARE_URL")
        or os.getenv("ONEDRIVE_FOLDER_SHARE_URL")
        or ""
    ).strip()
    if not folder_share_url:
        # If not configured, treat as "no restriction" on folder access.
        return
    from microsoft_graph import resolve_share_url
    resolve_share_url(folder_share_url, access_token)


# Best-effort local auth cache used by the frontend (optional).
# This is NOT the source of truth for access control: generation still validates the delegated token.
_AUTH_CACHE: dict = {"access_token": "", "expires_at": "", "user_email": "", "user_name": ""}


def _auth_cache_path() -> str:
    storage_dir = (os.getenv("LOCAL_STORAGE_DIR") or "").strip() or "local_storage"
    os.makedirs(storage_dir, exist_ok=True)
    return os.path.join(storage_dir, "auth_cache.json")


def _load_auth_cache_from_disk() -> None:
    try:
        path = _auth_cache_path()
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f) or {}
            if isinstance(data, dict):
                for k in list(_AUTH_CACHE.keys()):
                    if k in data:
                        _AUTH_CACHE[k] = str(data.get(k) or "")
    except Exception:
        # Never break auth endpoints due to cache issues
        pass


def _save_auth_cache_to_disk() -> None:
    try:
        path = _auth_cache_path()
        with open(path, "w", encoding="utf-8") as f:
            json.dump(_AUTH_CACHE, f)
    except Exception:
        pass


@agent.on_rest_post("/auth/validate", AuthValidateRequest, AuthValidateResponse)
async def auth_validate_endpoint(ctx: Context, req: AuthValidateRequest) -> AuthValidateResponse:
    """
    Validate a delegated Microsoft access token and (optionally) ensure the user
    has access to the configured OneDrive folder (and/or is allowlisted).
    """
    user_email = ""
    try:
        ctx.logger.info("Received /auth/validate request")
        token = (req.access_token or "").strip()
        if not token:
            ctx.logger.warning("Missing access_token in /auth/validate request")
            return AuthValidateResponse(
                timestamp=int(time.time()),
                message="Missing access_token",
                agent_address=ctx.agent.address,
                allowed=False,
                user_email="",
            )

        # Check if this is a simple login token
        if token.startswith("hco_token_"):
            ctx.logger.info("Simple login token detected")
            try:
                parts = token.split("_")
                if len(parts) >= 4:
                    token_timestamp = int(parts[2])
                    username = "_".join(parts[3:])
                    
                    if int(time.time()) - token_timestamp > (24 * 3600):
                        return AuthValidateResponse(
                            timestamp=int(time.time()),
                            message="Token expired",
                            agent_address=ctx.agent.address,
                            allowed=False,
                            user_email="",
                        )
                    
                    ctx.logger.info(f"✅ Simple login token valid for {username}")
                    return AuthValidateResponse(
                        timestamp=int(time.time()),
                        message="OK",
                        agent_address=ctx.agent.address,
                        allowed=True,
                        user_email=username,
                    )
            except Exception as e:
                ctx.logger.warning(f"Invalid simple login token: {str(e)}")
                return AuthValidateResponse(
                    timestamp=int(time.time()),
                    message="Invalid token",
                    agent_address=ctx.agent.address,
                    allowed=False,
                    user_email="",
                )

        # Microsoft token: Drive access gate first: user must be able to resolve the configured folder share URL.
        try:
            ctx.logger.info("Checking OneDrive folder access")
            _check_user_can_access_onedrive_folder(token)
            ctx.logger.info("✅ OneDrive folder access confirmed")
        except Exception as e:
            ctx.logger.warning(f"OneDrive folder access check failed: {str(e)}")
            # uAgents enforces schema validity; never raise here.
            return AuthValidateResponse(
                timestamp=int(time.time()),
                message=f"Access denied: {str(e)}",
                agent_address=ctx.agent.address,
                allowed=False,
                user_email="",
            )

        try:
            ctx.logger.info("Getting user email from token")
            email = _graph_me_email(token)
            user_email = email
            ctx.logger.info(f"✅ User email: {email}")
        except Exception as e:
            ctx.logger.warning(f"Could not get user email: {str(e)}")
            email = ""
            user_email = ""
        
        allowed_emails = _parse_csv_env_list("HCO_ALLOWED_LOGIN_EMAILS")
        if allowed_emails and (not email or email.lower() not in allowed_emails):
            ctx.logger.warning(f"User {email} not in allowed list: {allowed_emails}")
            return AuthValidateResponse(
                timestamp=int(time.time()),
                message="Access denied: your Microsoft account is not allowed.",
                agent_address=ctx.agent.address,
                allowed=False,
                user_email=user_email,
            )

        ctx.logger.info(f"✅ Auth validation successful for {user_email}")
        return AuthValidateResponse(
            timestamp=int(time.time()),
            message="OK",
            agent_address=ctx.agent.address,
            allowed=True,
            user_email=user_email,
        )
    except Exception as e:
        ctx.logger.error(f"Unexpected error in /auth/validate: {str(e)}", exc_info=True)
        return AuthValidateResponse(
            timestamp=int(time.time()),
            message=f"Access denied: {str(e)}",
            agent_address=ctx.agent.address,
            allowed=False,
            user_email=user_email,
        )


@agent.on_rest_post("/auth/login", LoginRequest, LoginResponse)
async def login_endpoint(ctx: Context, req: LoginRequest) -> LoginResponse:
    """Simple username/password login with hardcoded credentials."""
    ctx.logger.info(f"Received login request for username: {req.username}")
    
    HARDCODED_USERNAME = "hcoadmin"
    HARDCODED_PASSWORD = "hcoadmin123"
    
    if req.username == HARDCODED_USERNAME and req.password == HARDCODED_PASSWORD:
        token = f"hco_token_{int(time.time())}_{req.username}"
        expires_at = int(time.time()) + (24 * 3600)
        
        ctx.logger.info(f"✅ Login successful for {req.username}")
        return LoginResponse(
            timestamp=int(time.time()),
            message="Login successful",
            agent_address=ctx.agent.address,
            token=token,
            expires_at=expires_at,
            user_email=req.username,
            allowed=True,
        )
    else:
        ctx.logger.warning(f"❌ Login failed for {req.username}")
        return LoginResponse(
            timestamp=int(time.time()),
            message="Invalid username or password",
            agent_address=ctx.agent.address,
            token="",
            expires_at=0,
            user_email="",
            allowed=False,
        )


@agent.on_rest_post("/auth/status", AuthStatusRequest, AuthStatusResponse)
async def auth_status_endpoint(ctx: Context, req: AuthStatusRequest) -> AuthStatusResponse:
    _load_auth_cache_from_disk()
    authenticated = bool((_AUTH_CACHE.get("access_token") or "").strip())
    return AuthStatusResponse(
        timestamp=int(time.time()),
        agent_address=ctx.agent.address,
        authenticated=authenticated,
        expires_at=str(_AUTH_CACHE.get("expires_at") or ""),
    )


@agent.on_rest_post("/auth/sync", AuthSyncRequest, AuthSyncResponse)
async def auth_sync_endpoint(ctx: Context, req: AuthSyncRequest) -> AuthSyncResponse:
    _AUTH_CACHE.update(
        {
            "access_token": (req.access_token or "").strip(),
            "expires_at": (req.expires_at or "").strip(),
            "user_email": (req.user_email or "").strip(),
            "user_name": (req.user_name or "").strip(),
        }
    )
    _save_auth_cache_to_disk()
    return AuthSyncResponse(
        timestamp=int(time.time()),
        agent_address=ctx.agent.address,
        ok=True,
        message="OK",
    )


@agent.on_rest_post("/auth/clear", AuthClearRequest, AuthClearResponse)
async def auth_clear_endpoint(ctx: Context, req: AuthClearRequest) -> AuthClearResponse:
    _AUTH_CACHE.update({"access_token": "", "expires_at": "", "user_email": "", "user_name": ""})
    _save_auth_cache_to_disk()
    return AuthClearResponse(
        timestamp=int(time.time()),
        agent_address=ctx.agent.address,
        ok=True,
        message="OK",
    )

class CertificateQueryRequest(Model):
    query: str
    
class CertificateQueryResponse(Model):
    timestamp: int
    message: str
    agent_address: str
    query_type: str  # "verification", "download", "inquiry"
    certificate_no: Optional[str] = None
    download_url: Optional[str] = None
    processed: bool
    filename: Optional[str] = None
    found: bool = False
    verified_product_names: List[str] = []
    verified_product_codes: List[str] = []
    missing_product_names: List[str] = []
    missing_product_codes: List[str] = []


# Storage configuration
STORAGE_URL = os.getenv("AGENTVERSE_URL", "https://agentverse.ai") + "/v1/storage"

# ASI:one configuration (commented out - using OpenAI instead)
# ASI_ONE_API_KEY = os.getenv("ASI_ONE_API_KEY", "YOUR_ASI_ONE_API_KEY")
# ASI_ONE_BASE_URL = "https://api.asi1.ai/v1"

# OpenAI configuration
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "YOUR_OPENAI_API_KEY")

# Initialize OpenAI client (replacing ASI:one)
# asi_client = OpenAI(
#     api_key=ASI_ONE_API_KEY,
#     base_url=ASI_ONE_BASE_URL
# )

# Initialize OpenAI client
openai_client = OpenAI(
    api_key=OPENAI_API_KEY
)

# Certificate generation configuration
UPLOAD_FOLDER = 'generated_certificates'
CSV_FILE = 'certificates.csv'
TEMPLATE_FILE = 'template.png'

# Create directories if they don't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Create the chat protocol
chat_proto = Protocol(spec=chat_protocol_spec)


def create_text_chat(text: str) -> ChatMessage:
    return ChatMessage(
        timestamp=datetime.now(timezone.utc),
        msg_id=uuid4(),
        content=[TextContent(type="text", text=text)],
    )

def create_metadata(metadata: dict[str, str]) -> ChatMessage:
    return ChatMessage(
        timestamp=datetime.now(timezone.utc),
        msg_id=uuid4(),
        content=[MetadataContent(
            type="metadata",
            metadata=metadata,
        )],
    )

def initialize_csv():
    """Initialize CSV file with headers if it doesn't exist"""
    if not os.path.exists(CSV_FILE):
        with open(CSV_FILE, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow([
                'certificate_no',
                'issue_date', 
                'company_reg_no',
                'company_name'
            ])

def save_to_sheets(data, png_filename, pdf_filename):
    """
    Legacy hook (Google Sheets) removed.
    Keep the function to avoid touching older call sites, but no-op it.
    """
    return False

def extract_certificate_number_from_text(text):
    """Extract certificate number from text message using OpenAI"""
    if not text or OPENAI_API_KEY == "YOUR_OPENAI_API_KEY":
        # Fallback to regex patterns if OpenAI is not configured
        return extract_certificate_number_regex_fallback(text)
    
    try:
        prompt = f"""
Extract the COMPLETE certificate number from the following text. Certificate numbers may contain letters, numbers, slashes (/), and dashes (-). 

IMPORTANT: Keep ALL parts of the certificate number including slashes and dashes. Do NOT remove or truncate any part.

Examples of valid certificate numbers:
- HCO/RAO/TEST/202522323233 (keep ALL parts with slashes)
- HCO-2024-001 (keep dashes)
- HCO/RAO/091024 (keep all slashes)
- ABC123DEF456  
- DEF20240001
- XYZ/2024/123
- iwhfoihwe
- simple123
- test456

Text: "{text}"

Be EXTREMELY liberal in what you consider a certificate number, but ALWAYS include the COMPLETE identifier with ALL slashes, dashes, and parts.

Examples:
- "download certificate HCO/RAO/TEST/202522323233" -> extract "HCO/RAO/TEST/202522323233" (keep EVERYTHING)
- "i want to download certificate iwhfoihwe" -> extract "iwhfoihwe"
- "download ABC123" -> extract "ABC123"  
- "get certificate test" -> extract "test"
- "certificate number HCO/RAO/091024" -> extract "HCO/RAO/091024" (keep slashes)
- "verify HCO/RAO/TEST/202522323233" -> extract "HCO/RAO/TEST/202522323233" (COMPLETE number)

If you find ANY potential certificate identifier, return only that COMPLETE identifier without any explanation.
If you absolutely cannot find anything that could be a certificate number, return "NONE".
"""

        response = openai_client.chat.completions.create(
            model="gpt-5.2", 
            messages=[
                {"role": "system", "content": "You are an expert at extracting certificate numbers from text. Return only the certificate number or 'NONE'."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1,
            max_completion_tokens=50
        )
        
        result = (response.choices[0].message.content or "").strip()
        
        if result == "NONE" or not result:
            return None
        
        return result
        
    except Exception as e:
        hlog.warn("VERIFY", "openai cert extraction failed", reason=str(e))
        # Fallback to regex patterns
        return extract_certificate_number_regex_fallback(text)


# ============================================
# Product verification helpers
# ============================================

def _split_csv(raw: Any) -> List[str]:
    """Split a comma-separated string into a trimmed list, ignoring empties."""
    if not raw:
        return []
    return [s.strip() for s in str(raw).split(",") if s.strip()]


def _get_row_col(row: dict, *candidate_keys: str) -> Any:
    """Return the value for the first key that exists in *row* (case-insensitive)."""
    lower_row = {k.strip().lower(): v for k, v in row.items()}
    for key in candidate_keys:
        val = lower_row.get(key.strip().lower())
        if val is not None:
            return val
    return None


def _llm_match_products(user_items: List[str], excel_items: List[str], item_type: str) -> Dict[str, bool]:
    """Use quick exact match first, then LLM for fuzzy matching of remaining items."""
    results: Dict[str, bool] = {}

    # Quick case-insensitive exact match
    excel_lower = {e.strip().lower() for e in excel_items}
    unresolved: List[str] = []
    for item in user_items:
        if item.strip().lower() in excel_lower:
            results[item] = True
        else:
            unresolved.append(item)

    if not unresolved:
        hlog.info("VERIFY", "products quick match", resolved=len(results), total=len(user_items))
        return results

    # LLM fuzzy match for remaining items
    try:
        prompt = (
            f"You are a product verification assistant. Compare user-provided product {item_type} "
            f"against the certificate's product {item_type} list. For each user item, determine if it "
            f"matches any certificate item (consider casing, dashes, spacing, abbreviations).\n\n"
            f"Certificate {item_type}: {json.dumps(excel_items)}\n"
            f"User {item_type} to verify: {json.dumps(unresolved)}\n\n"
            f"Return a JSON object mapping each user item to true (match found) or false (no match).\n"
            f"Example: {{\"Beef-XP 1.8kg\": true, \"Unknown Product\": false}}"
        )

        hlog.info("VERIFY", "products llm match", unresolved=len(unresolved), item_type=item_type)
        resp = _openai_chat_completion_with_retry(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
            max_tokens=512,
        )
        raw = (resp.choices[0].message.content or "").strip()
        if raw.startswith("```"):
            raw = re.sub(r"^```(?:json)?\s*", "", raw)
            raw = re.sub(r"\s*```$", "", raw)
        llm_result = json.loads(raw)
        if isinstance(llm_result, dict):
            for item in unresolved:
                results[item] = bool(llm_result.get(item, False))
        else:
            hlog.warn("VERIFY", "products llm returned non-dict", got=str(type(llm_result)))
            for item in unresolved:
                results[item] = False
    except Exception as e:
        hlog.warn("VERIFY", "products llm error", reason=str(e))
        for item in unresolved:
            results[item] = False

    return results


def _verify_products_against_excel(
    certificate_no: str,
    product_names: List[str] | None = None,
    product_codes: List[str] | None = None,
) -> tuple:
    """Return (ok, payload) where payload matches ProductVerifyResponse fields (minus timestamp/agent_address)."""
    _empty_payload: Dict[str, Any] = {
        "certificate_found": False,
        "verified": False,
        "download_url": "",
        "verified_product_names": [],
        "verified_product_codes": [],
        "missing_product_names": [],
        "missing_product_codes": [],
    }

    cert_no = (certificate_no or "").strip()
    if not cert_no:
        return True, {
            "certificate_no": "",
            **_empty_payload,
            "message": "No certificate number provided.",
        }

    ms_excel_share_url = os.getenv("HCO_EXCEL_SHARE_URL") or os.getenv("EXCEL_SHARE_URL")
    ms_table_name = (
        os.getenv("HCO_EXCEL_TABLE_NAME")
        or os.getenv("EXCEL_TABLE_NAME")
        or "Certificates"
    )

    if not ms_excel_share_url:
        return True, {
            "certificate_no": cert_no,
            **_empty_payload,
            "message": "Product verification is not configured.",
        }

    try:
        from microsoft_graph import get_access_token, find_row_in_excel_table_by_column_value
        token = get_access_token()
    except Exception as e:
        return True, {
            "certificate_no": cert_no,
            **_empty_payload,
            "message": f"Authentication error: {e}",
        }

    row = find_row_in_excel_table_by_column_value(
        ms_excel_share_url,
        ms_table_name,
        column_name="certificate_no",
        match_value=cert_no,
        token=token,
    )

    if not row:
        return True, {
            "certificate_no": cert_no,
            "certificate_found": False,
            "download_url": "",
            "verified": False,
            "message": f"Certificate {cert_no} not found.",
            **_empty_payload,
            "missing_product_names": list(product_names or []),
            "missing_product_codes": list(product_codes or []),
        }

    # Excel products_code and products_name are comma-separated; parse for this certificate
    products_code_raw = _get_row_col(row, "products_code", "Product Code", "Products Code", "product_code")
    products_name_raw = _get_row_col(row, "products_name", "Product Name", "Products Name", "product_name")
    existing_codes_list = _split_csv(products_code_raw)
    existing_names_list = _split_csv(products_name_raw)

    wanted_codes = [str(c).strip() for c in (product_codes or []) if str(c).strip()]
    wanted_names = [str(n).strip() for n in (product_names or []) if str(n).strip()]

    # Use LLM to match user items against Excel items (handles casing, dashes, spacing, etc.)
    code_results: Dict[str, bool] = {}
    if wanted_codes and existing_codes_list:
        code_results = _llm_match_products(wanted_codes, existing_codes_list, "codes")
    name_results: Dict[str, bool] = {}
    if wanted_names and existing_names_list:
        name_results = _llm_match_products(wanted_names, existing_names_list, "names")

    verified_codes = [c for c in wanted_codes if code_results.get(c, False)]
    missing_codes = [c for c in wanted_codes if not code_results.get(c, False)]
    verified_names = [n for n in wanted_names if name_results.get(n, False)]
    missing_names = [n for n in wanted_names if not name_results.get(n, False)]

    verified = (len(missing_codes) == 0) and (len(missing_names) == 0)
    dl = str(row.get("certificate_url") or "").strip()

    verified_count = len(verified_codes) + len(verified_names)
    missing_count = len(missing_codes) + len(missing_names)
    total_count = verified_count + missing_count

    # Build structured response
    msg_parts: List[str] = []
    status_icon = "✅" if verified else "⚠️"
    msg_parts.append(f"{status_icon} **Product Verification Report**")
    msg_parts.append(f"**Certificate:** {cert_no}")
    msg_parts.append(f"**Total products checked:** {total_count}")
    msg_parts.append("")

    if verified_count > 0:
        msg_parts.append(f"**Verified ({verified_count}/{total_count}):**")
        for name in verified_names:
            msg_parts.append(f"  - {name} (name)")
        for code in verified_codes:
            msg_parts.append(f"  - {code} (code)")

    if missing_count > 0:
        if verified_count > 0:
            msg_parts.append("")
        msg_parts.append(f"**Not Found ({missing_count}/{total_count}):**")
        for name in missing_names:
            msg_parts.append(f"  - {name} (name)")
        for code in missing_codes:
            msg_parts.append(f"  - {code} (code)")

    if total_count == 0:
        msg_parts.append("No product names or codes were provided to verify.")

    msg_parts.append("")
    if verified:
        msg_parts.append("**Result:** All products are verified on this certificate.")
    elif missing_count > 0 and verified_count > 0:
        msg_parts.append(f"**Result:** {missing_count} product(s) could not be found on this certificate. Please double-check the spelling or contact HCO for assistance.")
    elif missing_count > 0:
        msg_parts.append("**Result:** None of the products were found on this certificate. Please verify the certificate number and product details, or contact HCO for assistance.")

    msg = "\n".join(msg_parts)

    return True, {
        "certificate_no": cert_no,
        "certificate_found": True,
        "download_url": dl,
        "verified": verified,
        "verified_product_names": verified_names,
        "verified_product_codes": verified_codes,
        "missing_product_names": missing_names,
        "missing_product_codes": missing_codes,
        "message": msg,
    }


def _extract_products_from_text(text: str) -> tuple:
    """Extract product names and codes from a free-text query.
    Uses regex patterns first, then falls back to LLM for natural language.
    Returns (names: List[str], codes: List[str])."""
    codes: List[str] = []
    names: List[str] = []

    # --- Pass 1: structured regex patterns ---
    m_codes = re.search(r"product\s*codes?\s*[:\-]\s*([^\n]+)", text, flags=re.IGNORECASE)
    if m_codes:
        codes = [c.strip() for c in m_codes.group(1).split(",") if c.strip()]

    m_names = re.search(r"product\s*names?\s*[:\-]\s*([^\n]+)", text, flags=re.IGNORECASE)
    if m_names:
        names = [n.strip() for n in m_names.group(1).split(",") if n.strip()]

    if not codes:
        m_codes2 = re.search(r"codes?\s*[:\-]\s*([^\n]+)", text, flags=re.IGNORECASE)
        if m_codes2:
            codes = [c.strip() for c in m_codes2.group(1).split(",") if c.strip()]

    if not names:
        m_names2 = re.search(r"names?\s*[:\-]\s*([^\n]+)", text, flags=re.IGNORECASE)
        if m_names2:
            names = [n.strip() for n in m_names2.group(1).split(",") if n.strip()]

    if not names:
        quoted = re.findall(r"product\s*names?\s*\"([^\"]+)\"", text, flags=re.IGNORECASE)
        if not quoted:
            quoted = re.findall(r"product\s*name\s*\"([^\"]+)\"", text, flags=re.IGNORECASE)
        if not quoted:
            quoted = re.findall(r"product\s*\"([^\"]+)\"", text, flags=re.IGNORECASE)
        if not quoted:
            quoted = re.findall(r"\"([^\"]+)\"", text)
        if quoted:
            names = [q.strip() for q in quoted if q and q.strip()]

    if not codes:
        quoted_codes = re.findall(r"product\s*codes?\s*\"([^\"]+)\"", text, flags=re.IGNORECASE)
        if not quoted_codes:
            quoted_codes = re.findall(r"product\s*code\s*\"([^\"]+)\"", text, flags=re.IGNORECASE)
        if quoted_codes:
            codes = [q.strip() for q in quoted_codes if q and q.strip()]

    # --- Pass 2: LLM fallback for natural language ---
    if not names and not codes:
        try:
            cert_no = extract_certificate_number_from_text(text) or ""
            clean_text = text
            if cert_no:
                clean_text = clean_text.replace(cert_no, "").strip()

            prompt = (
                "Extract product names and product codes from the following user query. "
                "The user is asking to verify products against a certificate.\n\n"
                "Rules:\n"
                "- Product NAMES are descriptive labels like 'Beef-XP 1.8kg', 'Chicken Nuggets 500g', 'BCAA 450g'\n"
                "- Product CODES are short alphanumeric identifiers like 'PRD001', 'BF-100', 'CK200'\n"
                "- Do NOT include certificate numbers, verbs, or filler words\n"
                "- If ambiguous whether something is a name or code, treat it as a name\n\n"
                f"Query: \"{clean_text}\"\n\n"
                "Respond with ONLY a JSON object like:\n"
                '{"names": ["Product A", "Product B"], "codes": ["PRD001"]}\n'
                "If no products found, return: {\"names\": [], \"codes\": []}"
            )

            resp = _openai_chat_completion_with_retry(
                model="gpt-4o-mini",
                messages=[{"role": "user", "content": prompt}],
                temperature=0,
                max_tokens=512,
            )
            raw = (resp.choices[0].message.content or "").strip()
            if raw.startswith("```"):
                raw = re.sub(r"^```(?:json)?\s*", "", raw)
                raw = re.sub(r"\s*```$", "", raw)
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                llm_names = [str(n).strip() for n in (parsed.get("names") or []) if str(n).strip()]
                llm_codes = [str(c).strip() for c in (parsed.get("codes") or []) if str(c).strip()]
                if llm_names or llm_codes:
                    names = llm_names
                    codes = llm_codes
        except Exception as e:
            hlog.warn("VERIFY", "product extract llm failed", reason=str(e))

    return names, codes


def _do_product_verification(certificate_no: str, text: str) -> dict:
    """Run product verification and return a dict with 'message', 'verified', etc."""
    names, codes = _extract_products_from_text(text)

    ok, payload = _verify_products_against_excel(
        certificate_no=certificate_no,
        product_names=names,
        product_codes=codes,
    )
    return payload


def _process_text_query(query: str) -> dict:
    """Common text query processor shared by /chat, /certificate/query and chat protocol.

    Returns a dict with:
        message, query_type, processed, certificate_no, certificate_found,
        download_url, filename
    """
    result: dict = {
        "message": "",
        "query_type": "unknown",
        "processed": False,
        "certificate_no": None,
        "certificate_found": False,
        "download_url": None,
        "filename": None,
    }

    if not query or not query.strip():
        result["message"] = "Please provide a query."
        return result

    q_lower = query.lower()
    certificate_no = extract_certificate_number_from_text(query)
    result["certificate_no"] = certificate_no

    # Product verification keyword check (most specific match first)
    has_product_kw = ("product" in q_lower or "products" in q_lower)
    has_verify_action = any(w in q_lower for w in (
        "verify", "validate", "check", "available", "listed", "present",
        "find", "exist", "confirm", "match",
    ))
    wants_product_verify = has_product_kw and has_verify_action

    query_type = classify_query(query)
    result["query_type"] = query_type

    # --- product verification ---
    if wants_product_verify or query_type == "product_verification":
        result["query_type"] = "product_verification"
        if certificate_no:
            try:
                payload = _do_product_verification(certificate_no, query)
                result["message"] = str(payload.get("message") or "")
                result["processed"] = bool(payload.get("verified"))
                result["certificate_found"] = bool(payload.get("certificate_found"))
                result["download_url"] = str(payload.get("download_url") or "").strip() or None
                result["verified_product_names"] = list(payload.get("verified_product_names") or [])
                result["verified_product_codes"] = list(payload.get("verified_product_codes") or [])
                result["missing_product_names"] = list(payload.get("missing_product_names") or [])
                result["missing_product_codes"] = list(payload.get("missing_product_codes") or [])
                return result
            except Exception as e:
                hlog.warn("VERIFY", "product verify text query failed", reason=str(e))
                result["message"] = (
                    f"Error verifying products for certificate {certificate_no}. "
                    "Please try again or use the Product Verification form for more reliable results."
                )
                return result
        else:
            result["message"] = (
                "Please provide a certificate number along with the product names or codes to verify.\n\n"
                "**Example:** _verify product names Beef-XP, Chicken 500g for certificate HCO-2024-001_"
            )
            return result

    # --- inquiry ---
    if query_type == "inquiry":
        result["query_type"] = "inquiry"
        try:
            search_results = search_hco_website(query)
            final_answer = generate_final_answer(query, search_results)
            result["message"] = final_answer
            result["processed"] = True
        except Exception:
            result["message"] = (
                "I'm having trouble processing your inquiry right now. "
                "Please contact HCO directly at info@hcoltd.co.uk or +44 (0) 333 577 0902 for assistance."
            )
        return result

    # --- marketing ---
    if query_type == "marketing":
        result["query_type"] = "marketing"
        try:
            marketing_content = generate_marketing_content(query)
            result["message"] = marketing_content
            result["processed"] = True
        except Exception:
            result["message"] = (
                "I'm having trouble generating marketing content right now. "
                "For marketing materials and promotional content, please contact HCO directly at "
                "info@hcoltd.co.uk or +44 (0) 333 577 0902."
            )
        return result

    # --- generation ---
    if query_type == "generation":
        result["query_type"] = "generation"
        result["message"] = (
            "🔒 **Certificate Generation Available**\n\n"
            "To generate a new certificate, please use the step-by-step certificate generation interface. "
            "Simply type **'generate certificate'** in the chat to start the guided process.\n\n"
            "For assistance with certificate generation, contact HCO at info@hcoltd.co.uk or +44 (0) 333 577 0902."
        )
        result["processed"] = True
        return result

    # --- download ---
    if query_type == "download":
        result["query_type"] = "download"
        if certificate_no:
            try:
                from database import get_certificate_from_db, get_certificate_file_from_db

                certificate_data = get_certificate_from_db(certificate_no)
                if certificate_data:
                    file_data = get_certificate_file_from_db(certificate_no, "pdf")
                    file_type = "pdf"
                    if not file_data:
                        file_data = get_certificate_file_from_db(certificate_no, "png")
                        file_type = "png"

                    if file_data:
                        filename = f"{certificate_no}.{file_type}"
                        base_url = (
                            os.getenv("AGENT_URL") or os.getenv("BASE_URL") or "http://localhost:8025"
                        ).rstrip("/")
                        download_url = f"{base_url}/download-certificate"

                        result["message"] = (
                            f"✅ **Certificate Found!**\n\n"
                            f"📋 **Certificate Number:** {certificate_no}\n"
                            f"📄 **File:** {filename}\n"
                            f"💾 **Status:** Ready for download"
                        )
                        result["processed"] = True
                        result["certificate_found"] = True
                        result["download_url"] = download_url
                        result["filename"] = filename
                    else:
                        result["message"] = f"Certificate {certificate_no} found but no file data available for download."
                else:
                    result["message"] = f"Certificate {certificate_no} not found in database."
            except Exception as e:
                result["message"] = f"Error processing download: {e}"
        else:
            result["message"] = "Please provide a certificate number to download."
        return result

    # --- default: certificate verification ---
    result["query_type"] = "verification"
    if certificate_no:
        is_valid, certificate_data = verify_certificate(certificate_no)
        if is_valid:
            company_name = certificate_data.get("company_name", "Unknown Company")
            issue_date = certificate_data.get("issue_date", "Unknown Date")
            result["message"] = (
                f"✅ **Certificate Verified!**\n\n"
                f"This is a valid HCO certificate for **{company_name}**, issued on **{issue_date}**.\n\n"
                f"Certificate Number: {certificate_no}"
            )
            result["processed"] = True
            result["certificate_found"] = True
        else:
            result["message"] = (
                "❌ **Certificate Not Valid**\n\n"
                "This certificate is not valid. If you need a valid HCO certificate, please apply at "
                "https://www.hcoltd.co.uk/registration or contact HCO directly for assistance at "
                "info@hcoltd.co.uk or +44 (0) 333 577 0902."
            )
    else:
        result["message"] = (
            "I couldn't find a certificate number in your message. Please provide either:\n"
            "• An image of the certificate to analyze\n"
            "• A text message containing the certificate number (e.g., 'HCO-2024-001')"
        )
    return result


def classify_query(text):
    """Classify query as 'inquiry', 'verification', 'product_verification', 'download', or 'marketing' using OpenAI"""
    if not text:
        return "inquiry"  # Default to inquiry if no text
    
    text_lower = text.lower()

    # Check for product verification FIRST (before generic verification)
    has_product_kw = ("product" in text_lower or "products" in text_lower)
    has_verify_kw = any(w in text_lower for w in (
        "verify", "validate", "check", "available", "listed", "present",
        "find", "exist", "confirm", "match",
    ))
    if has_product_kw and has_verify_kw:
        return "product_verification"
    
    # Check for marketing keywords
    marketing_keywords = [
        "marketing", "marketing content", "generate marketing", "create marketing",
        "marketing material", "promotional content", "advertisement", "promote hco",
        "marketing copy", "content for marketing", "hco marketing", "marketing text",
        "promotional text", "advertising content", "marketing campaign", "campaign",
        "copywriting", "copy writing", "copy", "sales copy", "ad copy", "landing page",
        "brochure", "flyer", "poster", "tagline", "slogan", "headline",
        "social post", "social media", "linkedin", "instagram", "facebook", "twitter",
        "caption", "email", "newsletter", "press release", "product description",
        "website content", "seo", "google ads", "meta title", "meta description",
        "brand", "branding"
    ]
    
    # Check for generation keywords
    generation_keywords = [
        "generate certificate", "create certificate", "i want to generate certificate",
        "i need to generate certificate", "generate a certificate", "create a certificate",
        "make certificate", "i want to create certificate", "generate new certificate"
    ]
    
    # Check for download keywords
    download_keywords = [
        "download", "get certificate", "retrieve certificate", "get copy", "download copy", 
        "i want to download", "i need to download", "want certificate", "need certificate"
    ]
    
    # Check for verification keywords
    verification_keywords = [
        "verify", "check certificate", "validate", "certificate number", "cert no", "certificate id",
        "is valid", "authentication", "confirm certificate", "certificate status"
    ]
    
    # Look for certificate number patterns
    certificate_patterns = [
        r'\b[A-Z]{2,4}/[A-Z]{2,4}/\d{6}\b',  # HCO/RAO/091024
        r'\b[A-Z]{2,4}[-/]\d{4}[-/]\d{2,4}\b',  # HCO-2024-001
        r'\b[A-Z]{2,4}\d{4,8}\b',  # HCO20240001
        r'\b\d{4,8}[-/]\d{2,4}\b',  # 2024-001
        r'\b[A-Z]+\d+[A-Z]*\d*\b',  # ABC123DEF456
    ]
    
    # Check if query contains generation keywords
    has_generation_keywords = any(keyword in text_lower for keyword in generation_keywords)
    
    # Check if query contains marketing keywords
    has_marketing_keywords = any(keyword in text_lower for keyword in marketing_keywords)
    
    # Check if query contains download keywords
    has_download_keywords = any(keyword in text_lower for keyword in download_keywords)
    
    # Check if query contains verification keywords or certificate patterns
    has_verification_keywords = any(keyword in text_lower for keyword in verification_keywords)
    has_certificate_pattern = any(re.search(pattern, text) for pattern in certificate_patterns)
    
    if has_generation_keywords:
        return "generation"
    elif has_marketing_keywords:
        return "marketing"
    elif has_download_keywords:
        return "download"
    elif has_verification_keywords or has_certificate_pattern:
        return "verification"
    
    # If OpenAI is not configured, default to inquiry for general questions
    if OPENAI_API_KEY == "YOUR_OPENAI_API_KEY":
        return "inquiry"
    
    try:
        prompt = f"""
Classify the following text query into one of these categories:

1. "inquiry" - General questions about HCO, services, processes, information, company details, etc.
2. "verification" - If the user wants to verify/check/validate a certificate with a certificate number/ID
3. "product_verification" - If the user wants to verify/check/validate specific PRODUCTS (by name or code) against a certificate
4. "download" - If the user wants to download/get/retrieve a certificate file
5. "generation" - If the user wants to generate/create/make a new certificate
6. "marketing" - If the user wants to generate marketing content, promotional material, or advertising copy for HCO

Text: "{text}"

Examples of INQUIRY:
- "What is the mission of HCO?"
- "What services does HCO provide?"
- "How do I get certified?"
- "Tell me about HCO company"

Examples of VERIFICATION:
- "Check certificate HCO-2024-001"
- "Is certificate ABC123 valid?"
- "Verify certificate number DEF456"

Examples of PRODUCT_VERIFICATION:
- "Verify product Beef-XP 1.8kg with certificate HCO-2024-001"
- "Check product code ABC123 for certificate HCO/TEST/001"
- "Validate products Beef-XP, BCAA 450g against certificate HCO-2024-001"

Examples of DOWNLOAD:
- "Download certificate HCO-2024-001"
- "Get certificate copy"
- "Retrieve my certificate"

Examples of GENERATION:
- "I want to generate certificate"
- "Generate certificate"
- "Create certificate"
- "I need to create a certificate"
- "Generate a new certificate"
- "Make a certificate"

Examples of MARKETING:
- "Generate marketing content for HCO"
- "Create promotional material for HCO"
- "Generate marketing copy"
- "Create advertising content"
- "Generate promotional text for HCO services"
- "Write me a content for marketing campaign for HCO"
- "Create a marketing campaign for HCO"

Respond with only one word: "inquiry", "verification", "product_verification", "download", "generation", or "marketing"
"""

        response = openai_client.chat.completions.create(
            model="gpt-5.2",
            messages=[
                {"role": "system", "content": "You are an expert at classifying user queries into inquiry, verification, download, or marketing categories."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1,
            max_completion_tokens=10
        )
        
        result = (response.choices[0].message.content or "").strip().lower()
        
        if result in ["inquiry", "verification", "product_verification", "download", "generation", "marketing"]:
            return result
        else:
            return "inquiry"  # Default to inquiry if unclear
            
    except Exception as e:
        hlog.warn("APP", "query classify failed", reason=str(e))
        return "inquiry"

def get_relevant_urls(query):
    """Determine which HCO website URLs are relevant to the user's query"""
    query_lower = query.lower()
    
    # Define URL mappings with keywords
    url_mappings = {
        "https://www.hcoltd.co.uk/": ["general", "home", "main", "overview", "hco"],
        "https://www.hcoltd.co.uk/about": ["about", "company", "who", "what is", "background", "history", "mission", "vision"],
        "https://www.hcoltd.co.uk/certificationprocess": ["certification", "process", "how to", "procedure", "steps", "certify", "certificate process"],
        "https://www.hcoltd.co.uk/policies": ["policy", "policies", "terms", "conditions", "rules", "guidelines", "privacy"],
        "https://www.hcoltd.co.uk/faq": ["faq", "question", "questions", "help", "answer", "common", "frequently asked"],
        "https://www.hcoltd.co.uk/contact": ["contact", "phone", "email", "address", "reach", "get in touch", "location"],
        "https://www.hcoltd.co.uk/registration": ["register", "registration", "sign up", "apply", "application", "join"],
        "https://www.hcoltd.co.uk/services": ["service", "services", "offer", "provide", "what do", "offerings"],
        "https://www.hcoltd.co.uk/news": ["news", "updates", "announcement", "latest", "recent", "current"],
        "http://hcoltd.co.uk/blog": ["blog", "article", "post", "insights", "thoughts", "stories"],
        "https://www.hcoltd.co.uk/reviews": ["review", "reviews", "testimonial", "feedback", "rating", "experience"],
        "https://www.hcoltd.co.uk/jobs": ["job", "jobs", "career", "employment", "work", "hiring", "vacancy", "position"]
    }
    
    relevant_urls = []
    
    # Check each URL for keyword matches
    for url, keywords in url_mappings.items():
        for keyword in keywords:
            if keyword in query_lower:
                if url not in relevant_urls:
                    relevant_urls.append(url)
                break
    
    # If no specific matches, include general pages
    if not relevant_urls:
        relevant_urls = [
            "https://www.hcoltd.co.uk/",
            "https://www.hcoltd.co.uk/about",
            "https://www.hcoltd.co.uk/services"
        ]
    
    return relevant_urls

def scrape_website_content(url):
    """Scrape content from a single URL using requests and BeautifulSoup"""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Remove script and style elements
        for script in soup(["script", "style", "nav", "footer", "header"]):
            script.decompose()
        
        # Extract text content
        text_content = soup.get_text()
        
        # Clean up the text
        lines = (line.strip() for line in text_content.splitlines())
        chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
        text = ' '.join(chunk for chunk in chunks if chunk)
        
        return text
        
    except Exception as e:
        hlog.warn("APP", "scrape failed", url=url, reason=str(e))
        return ""

def search_hco_website(query):
    """Extract content from relevant HCO website pages using web scraping for inquiry-related questions"""
    relevant_urls = get_relevant_urls(query)
    hlog.info("APP", "scrape start", query=query, urls=len(relevant_urls))

    all_content = []

    try:
        for url in relevant_urls:
            try:
                content = scrape_website_content(url)

                if content and content.strip():
                    if len(content) > 2000:
                        content = content[:2000] + "...(content truncated)"

                    all_content.append(f"=== Content from {url} ===\n{content}\n")

                time.sleep(1)

            except Exception as url_error:
                hlog.warn("APP", "scrape failed", url=url, reason=str(url_error))
                continue

        if not all_content:
            return "No content could be extracted from any relevant HCO website pages."

        combined_content = "\n".join(all_content)
        if len(combined_content) > 8000:
            combined_content = combined_content[:8000] + "\n...(content truncated)"

        return combined_content

    except Exception as e:
        hlog.warn("APP", "scrape pipeline failed", reason=str(e))
        return f"Error occurred while scraping the HCO website: {str(e)}"

def get_relevant_links(query):
    """Get relevant HCO website links based on the user's query"""
    query_lower = query.lower()
    
    # Define links based on query content
    link_mappings = {
        # Registration/Application related
        ("register", "registration", "apply", "application", "sign up", "join", "get certified", "how to certify"): 
            "https://www.hcoltd.co.uk/registration",
        
        # Services related
        ("service", "services", "offer", "provide", "what do", "offerings"):
            "https://www.hcoltd.co.uk/services",
        
        # Certification process
        ("process", "procedure", "steps", "how to", "certification process", "certify"):
            "https://www.hcoltd.co.uk/certificationprocess",
        
        # About/Company info
        ("about", "company", "who", "mission", "vision", "background", "history"):
            "https://www.hcoltd.co.uk/about",
        
        # Contact related
        ("contact", "phone", "email", "address", "reach", "get in touch"):
            "https://www.hcoltd.co.uk/contact",
        
        # FAQ related
        ("question", "faq", "help", "common"):
            "https://www.hcoltd.co.uk/faq",
        
        # Policies
        ("policy", "policies", "terms", "conditions", "privacy"):
            "https://www.hcoltd.co.uk/policies",
        
        # Reviews/Testimonials
        ("review", "reviews", "testimonial", "feedback", "experience"):
            "https://www.hcoltd.co.uk/reviews",
        
        # Jobs/Careers
        ("job", "jobs", "career", "employment", "work", "hiring"):
            "https://www.hcoltd.co.uk/jobs",
        
        # News/Updates
        ("news", "updates", "latest", "announcement"):
            "https://www.hcoltd.co.uk/news"
    }
    
    relevant_links = []
    
    for keywords, link in link_mappings.items():
        if any(keyword in query_lower for keyword in keywords):
            if link not in relevant_links:
                relevant_links.append(link)
    
    return relevant_links

def generate_final_answer(query, search_results):
    """Generate final answer using OpenAI based on query and search results"""
    if not search_results or OPENAI_API_KEY == "your_openai_api_key_here":
        return "I couldn't find specific information about that. For detailed assistance, please contact HCO directly at info@hcoltd.co.uk or +44 (0) 333 577 0902."
    
    # Check if search results contain meaningful content
    if "No content could be extracted" in search_results or len(search_results.strip()) < 100:
        return "I don't have specific information about that right now. For accurate details, please reach out to HCO at info@hcoltd.co.uk or call +44 (0) 333 577 0902. They'll be happy to help you directly!"
    
    # Get relevant links for the query
    relevant_links = get_relevant_links(query)
    links_text = ""
    if relevant_links:
        links_text = f"\n\nRelevant links: {', '.join(relevant_links)}"
    
    try:
        prompt = f"""
You are a helpful AI assistant for HCO Ltd. Based on the information provided, answer the user's question naturally and conversationally.

User Query: "{query}"

Available Information:
{search_results}

Relevant Links to Include: {relevant_links}

Instructions:
1. Answer based ONLY on the information provided
2. Be conversational and helpful, like a friendly assistant who works for HCO
3. When relevant, include appropriate links in your response naturally (e.g., "you can register at https://www.hcoltd.co.uk/registration")
4. Never mention "website", "search results", "mentioned on the website", or any reference to data sources
5. Speak as if you personally know this information about HCO
6. If you can't answer fully, suggest contacting HCO directly at info@hcoltd.co.uk or +44 (0) 333 577 0902
7. Keep responses natural and human-like
8. Include relevant links when they would be helpful to the user
9. For registration/application queries, always mention the registration link
10. For process queries, include the certification process link
11. Avoid phrases like "according to", "based on", "as stated", "mentioned on"

Generate a natural, helpful response as if you're a knowledgeable HCO representative:
"""

        response = openai_client.chat.completions.create(
            model="gpt-5.2",
            messages=[
                {"role": "system", "content": "You are a knowledgeable HCO representative. Answer naturally as if you personally know the information. Never mention websites, data sources, or where information comes from. Include helpful links naturally."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3,
            max_completion_tokens=500
        )
        
        answer = (response.choices[0].message.content or "").strip()
        
        # Add links at the end if not already included in the response
        if relevant_links and not any(link in answer for link in relevant_links):
            answer += links_text
        
        return answer
        
    except Exception as e:
        hlog.warn("APP", "final answer llm failed", reason=str(e))
        return "I'm having trouble accessing that information right now. Please contact HCO directly at info@hcoltd.co.uk or +44 (0) 333 577 0902 for assistance!"

def generate_marketing_content(query):
    """Generate marketing content for HCO using OpenAI"""
    if OPENAI_API_KEY == "YOUR_OPENAI_API_KEY":
        return "Marketing content generation requires OpenAI API configuration. Please contact HCO directly at info@hcoltd.co.uk or +44 (0) 333 577 0902 for marketing materials."
    
    try:
        # First get some context about HCO from their website
        hco_context = """
HCO Ltd is a leading halal certification company that provides comprehensive halal certification services to food manufacturers, processors, and suppliers. 
They ensure products meet strict halal standards and requirements, helping businesses access global halal markets.
Their services include halal certification, auditing, consultation, and ongoing support for maintaining halal compliance.
"""

        prompt = f"""
You are a professional marketing copywriter specializing in halal certification and food industry marketing. 
Create compelling marketing content for HCO Ltd based on the user's request.

User Request: "{query}"

Context about HCO Ltd:
{hco_context}

Key messaging points to consider:
- Trust and reliability in halal certification
- Expertise and experience in the industry
- Global recognition and acceptance
- Comprehensive support throughout the certification process
- Helping businesses access lucrative halal markets
- Strict adherence to halal standards and requirements
- Professional and timely service delivery

Create marketing content that is:
1. Professional and credible
2. Engaging and persuasive
3. Focused on benefits to potential clients
4. Appropriate for the food and certification industry
5. Includes relevant contact information: info@hcoltd.co.uk or +44 (0) 333 577 0902
6. Mentions their website: https://www.hcoltd.co.uk

Keep the content concise but impactful (150-300 words).
"""

        response = _openai_chat_completion_with_retry(
            model="gpt-5.2",
            messages=[
                {
                    "role": "system",
                    "content": "You are an expert marketing copywriter specializing in halal certification and food industry marketing. Create professional, engaging marketing content.",
                },
                {"role": "user", "content": prompt},
            ],
            temperature=0.7,
            max_tokens=650,
        )
        
        marketing_content = (response.choices[0].message.content or "").strip()
        
        # Add a header to make it clear this is marketing content
        final_content = f"🎯 **Marketing Content for HCO Ltd**\n\n{marketing_content}"
        
        return final_content
        
    except Exception as e:
        hlog.warn("APP", "marketing content llm failed", reason=str(e))
        return "I'm having trouble generating marketing content right now. For marketing materials and promotional content, please contact HCO directly at info@hcoltd.co.uk or +44 (0) 333 577 0902. They'll be happy to provide you with professional marketing materials!"

def process_files_with_openai(files_data: List[Dict]) -> List[Dict]:
    """Process CSV/Excel files using OpenAI to extract product code, product name from final product names"""
    if OPENAI_API_KEY == "YOUR_OPENAI_API_KEY":
        return []
    
    extracted_products = []
    
    for file_data in files_data:
        try:
            filename = file_data.get('filename', 'unknown')
            file_content = file_data.get('content', '')
            
            hlog.info("GENERATE", "extract products start", file=filename)

            if filename.lower().endswith('.csv'):
                df = pd.read_csv(StringIO(file_content))
            elif filename.lower().endswith(('.xlsx', '.xls')):
                df = None
                excel_reading_success = False

                if not excel_reading_success:
                    try:
                        import openpyxl
                        import io
                        excel_file = io.BytesIO(file_content)
                        wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
                        sheet_names = wb.sheetnames

                        target_sheet_name = None
                        for sheet_name in sheet_names:
                            sheet_lower = sheet_name.lower().strip()
                            if sheet_lower == "final product name" or sheet_lower == "final product names":
                                target_sheet_name = sheet_name
                                break

                        if not target_sheet_name:
                            for sheet_name in sheet_names:
                                sheet_lower = sheet_name.lower()
                                if "final" in sheet_lower and "product" in sheet_lower and "name" in sheet_lower:
                                    target_sheet_name = sheet_name
                                    break

                        if not target_sheet_name:
                            for sheet_name in sheet_names:
                                sheet_lower = sheet_name.lower()
                                if "product" in sheet_lower or "final" in sheet_lower:
                                    target_sheet_name = sheet_name
                                    break

                        if not target_sheet_name:
                            target_sheet_name = sheet_names[0]

                        ws = wb[target_sheet_name]
                        data_rows = []
                        headers = []
                        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                            if row_idx == 0:
                                headers = [str(cell) if cell is not None else f"Column_{i}" for i, cell in enumerate(row)]
                            data_rows.append([str(cell) if cell is not None else "" for cell in row])

                        if data_rows:
                            df = pd.DataFrame(data_rows[1:], columns=headers) if len(data_rows) > 1 else pd.DataFrame(columns=headers)
                            excel_reading_success = True
                            hlog.info("GENERATE", "excel read", engine="openpyxl", sheet=target_sheet_name, rows=df.shape[0], cols=df.shape[1])

                    except Exception as openpyxl_error:
                        hlog.warn("GENERATE", "openpyxl failed", reason=str(openpyxl_error))

                if not excel_reading_success and filename.lower().endswith('.xls'):
                    try:
                        import xlrd
                        workbook = xlrd.open_workbook(file_contents=file_content)
                        sheet_names = workbook.sheet_names()

                        target_sheet = None
                        for sheet_name in sheet_names:
                            if "final" in sheet_name.lower() and "product" in sheet_name.lower():
                                target_sheet = workbook.sheet_by_name(sheet_name)
                                break

                        if not target_sheet:
                            target_sheet = workbook.sheet_by_index(0)

                        headers = [str(target_sheet.cell_value(0, col)) for col in range(target_sheet.ncols)]
                        data = []
                        for row in range(1, target_sheet.nrows):
                            data.append([str(target_sheet.cell_value(row, col)) for col in range(target_sheet.ncols)])

                        df = pd.DataFrame(data, columns=headers)
                        excel_reading_success = True
                        hlog.info("GENERATE", "excel read", engine="xlrd", rows=df.shape[0], cols=df.shape[1])

                    except Exception as xlrd_error:
                        hlog.warn("GENERATE", "xlrd failed", reason=str(xlrd_error))

                if not excel_reading_success:
                    hlog.warn("GENERATE", "excel unreadable, asking llm to synthesize sample products", file=filename)
                    
                    try:
                        # Create a mock analysis with OpenAI based on filename and common patterns
                        excel_prompt = f"""
You are analyzing an Excel file named "{filename}" that contains product information for halal certification.

Since the file appears to be corrupted or unreadable by traditional methods, please provide a sample extraction based on common patterns for this type of file.

Based on the filename and context, this appears to be a product list. Please generate a reasonable sample of products that might be in such a file, focusing on:
1. Product codes (like BMS7573, BMS8466, etc.)
2. Product names (like "Vitalife Blueberry & Kiwi", etc.)

Generate 5-10 sample products that would be typical for a halal certification product list.

Return in this exact JSON format:
{{
    "products": [
        {{
            "product_code": "sample_code",
            "product_name": "sample_product_name"
        }}
    ],
    "note": "Generated sample data due to file reading issues"
}}

Return only valid JSON, no explanations.
"""

                        response = openai_client.chat.completions.create(
                            model="gpt-5.2",
                            messages=[
                                {"role": "system", "content": "You are an expert at generating sample product data for halal certification. Return only valid JSON."},
                                {"role": "user", "content": excel_prompt}
                            ],
                            temperature=0.3,
                            max_completion_tokens=1000
                        )
                        
                        result_text = (response.choices[0].message.content or "").strip()
                        
                        # Parse the OpenAI response
                        import json
                        try:
                            cleaned_text = result_text.strip()
                            if cleaned_text.startswith('```json'):
                                cleaned_text = cleaned_text.replace('```json', '').replace('```', '').strip()
                            
                            result_json = json.loads(cleaned_text)
                            
                            if 'products' in result_json and isinstance(result_json['products'], list):
                                synth_count = 0
                                for product in result_json['products']:
                                    if 'product_code' in product and 'product_name' in product:
                                        extracted_products.append({
                                            'product_code': product['product_code'],
                                            'product_name': product['product_name'],
                                            'source_file': filename + " (OpenAI generated)"
                                        })
                                        synth_count += 1
                                hlog.warn("GENERATE", "synthesized products via llm", file=filename, count=synth_count)
                                continue
                            hlog.warn("GENERATE", "llm response had no products", file=filename)

                        except json.JSONDecodeError as json_error:
                            hlog.warn("GENERATE", "llm response not parseable, trying manual extract", file=filename, reason=str(json_error))
                            fallback_products = manual_extract_products(df, filename)
                            if fallback_products:
                                extracted_products.extend(fallback_products)
                                hlog.info("GENERATE", "manual extract", file=filename, count=len(fallback_products))

                    except Exception as openai_error:
                        hlog.warn("GENERATE", "llm analysis failed, trying manual extract", file=filename, reason=str(openai_error))
                        fallback_products = manual_extract_products(df, filename)
                        if fallback_products:
                            extracted_products.extend(fallback_products)
                            hlog.info("GENERATE", "manual extract", file=filename, count=len(fallback_products))
                        else:
                            extracted_products.append({
                                'product_code': 'PROCESSING_ERROR',
                                'product_name': f'Unable to extract products from {filename} - file format may not be supported',
                                'source_file': filename
                            })
                        continue

            else:
                hlog.warn("GENERATE", "unsupported file type", file=filename)
                continue
            
            # Focus specifically on the "Final Product Names" sheet since every file has this
            final_product_data = None
            sheet_names = []
            
            # If this is an Excel file, look specifically for "Final Product Names" sheet
            if filename.lower().endswith(('.xlsx', '.xls')):
                try:
                    import openpyxl
                    import io
                    import tempfile
                    
                    # Create temporary file for sheet analysis
                    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                        temp_file.write(file_content)
                        temp_file.flush()
                        temp_path = temp_file.name
                    
                    # Get all sheet names
                    wb = openpyxl.load_workbook(temp_path, read_only=True, data_only=True)
                    sheet_names = wb.sheetnames
                    wb.close()
                    
                    target_sheet = None
                    for sheet_name in sheet_names:
                        if "final" in sheet_name.lower() and "product" in sheet_name.lower():
                            target_sheet = sheet_name
                            break

                    if target_sheet:
                        try:
                            final_df = pd.read_excel(temp_path, sheet_name=target_sheet)
                            final_product_data = final_df.to_string(index=False)
                        except Exception as sheet_error:
                            hlog.warn("GENERATE", "sheet read failed", sheet=target_sheet, reason=str(sheet_error))
                            final_product_data = None
                    else:
                        if sheet_names:
                            target_sheet = sheet_names[0]
                            final_df = pd.read_excel(temp_path, sheet_name=target_sheet)
                            final_product_data = final_df.to_string(index=False)

                    import os
                    os.unlink(temp_path)

                except Exception as sheet_error:
                    hlog.warn("GENERATE", "sheet analysis failed", file=filename, reason=str(sheet_error))
                    final_product_data = df.head(20).to_string(index=False)
            else:
                # For non-Excel files, use the current dataframe
                final_product_data = df.head(20).to_string(index=False)
            
            # Create focused prompt for Final Product Names sheet
            if not final_product_data:
                final_product_data = "No Final Product Names sheet data available"
            
            prompt = f"""
You are analyzing the "Final Product Names" sheet from an Excel file named "{filename}" for halal certification. This sheet contains the definitive list of products that need to be certified.

FILENAME: {filename}
TARGET SHEET: Final Product Names
AVAILABLE SHEETS: {sheet_names}

FINAL PRODUCT NAMES SHEET DATA:
{final_product_data}

FOCUSED EXTRACTION RULES:
1. Focus ONLY on the "Final Product Names" sheet data above
2. Look for structured product information in table format
3. Extract product codes (like BMS7715, BMS7716, BMS7717, etc.)
4. Extract corresponding product names (like "DOLCETTO - Gingerbread Sugar Free Syrup")
5. Skip header rows, company information, and instruction text
6. Look for rows that contain both a product code AND a product name

EXPECTED PATTERNS IN FINAL PRODUCT NAMES SHEET:
- Product codes are typically in one column (often alphanumeric like BMS7715)
- Product names are in another column (descriptive text like "DOLCETTO - Hazelnut Sugar Free Syrup")
- May have additional columns for packaging details, additional information
- Skip rows with company name, address, and other header information
- Focus on the actual product listing table

EXTRACTION STRATEGY:
1. Identify the table structure in the Final Product Names sheet
2. Find the column with product codes (BMS, SKU, etc.)
3. Find the corresponding column with product names
4. Check for a "Packaging Details" or "Packaging" column and extract values if present
5. Extract each product code + name (+ packaging) from the table rows
6. Ignore header text and focus on the product data rows

Return data in this exact JSON format:
{{
    "products": [
        {{
            "product_code": "extracted_product_code",
            "product_name": "complete_product_name",
            "packaging_details": "packaging info if available, otherwise empty string"
        }}
    ]
}}

Return only valid JSON. Extract ALL product codes and names from the Final Product Names sheet.
"""

            response = openai_client.chat.completions.create(
                model="gpt-5.2",
                messages=[
                    {"role": "system", "content": "You are an expert at extracting product information from Excel files and analyzing all sheets. Return only valid JSON."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,
                max_completion_tokens=2000
            )
            
            result_text = (response.choices[0].message.content or "").strip()
            
            # Parse JSON response
            import json
            try:
                # Clean the response text
                cleaned_text = result_text.strip()
                if cleaned_text.startswith('```json'):
                    cleaned_text = cleaned_text.replace('```json', '').replace('```', '').strip()
                
                result_json = json.loads(cleaned_text)
                
                if 'products' in result_json and isinstance(result_json['products'], list):
                    valid_count = 0
                    for product in result_json['products']:
                        if 'product_code' in product and 'product_name' in product:
                            extracted_products.append({
                                'product_code': product['product_code'],
                                'product_name': product['product_name'],
                                'source_file': filename,
                            })
                            valid_count += 1
                    hlog.info("GENERATE", "products extracted", source="llm", file=filename, count=valid_count)
                    continue
                else:
                    hlog.warn("GENERATE", "llm returned invalid structure, using manual extract", file=filename)
                    fallback_products = manual_extract_products(df, filename)
                    extracted_products.extend(fallback_products)
                    continue

            except json.JSONDecodeError as json_error:
                hlog.warn("GENERATE", "llm response not parseable, using manual extract", file=filename, reason=str(json_error))
                fallback_products = manual_extract_products(df, filename)
                extracted_products.extend(fallback_products)
                continue

        except Exception as e:
            hlog.warn("GENERATE", "csv processing failed", file=filename, reason=str(e))
            continue
    
    return extracted_products

def manual_extract_products(df: pd.DataFrame, filename: str) -> List[Dict]:
    """Fallback manual extraction when OpenAI fails"""
    products = []
    
    try:
        # Check if this is the VITALIFEHCO format (has 'Unnamed' columns)
        is_vitalife_format = any('Unnamed:' in str(col) for col in df.columns)
        
        if is_vitalife_format:
            hlog.info("GENERATE", "vitalife format detected", file=filename)
            for index, row in df.iterrows():
                try:
                    col1_val = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ""
                    col2_val = str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else ""

                    if (col1_val and col1_val not in ['nan', 'Product Code'] and
                        col2_val and 'vitalife' in col2_val.lower()):

                        products.append({
                            'product_code': col1_val,
                            'product_name': col2_val,
                            'source_file': filename
                        })
                except Exception:
                    continue
        else:
            # Standard extraction logic for regular Excel/CSV files
            code_columns = []
            name_columns = []
            packaging_columns = []
            
            for col in df.columns:
                col_lower = str(col).lower()
                if any(keyword in col_lower for keyword in ['sku', 'code', 'id', 'item']):
                    code_columns.append(col)
                if any(keyword in col_lower for keyword in ['name', 'product', 'description', 'title']):
                    name_columns.append(col)
                if any(keyword in col_lower for keyword in ['packaging', 'packing', 'pack detail']):
                    packaging_columns.append(col)
            
            # Prefer specific columns
            code_col = None
            name_col = None
            packaging_col = None
            
            # Select best code column
            if code_columns:
                for col in code_columns:
                    if 'product' in str(col).lower() or 'sku' in str(col).lower():
                        code_col = col
                        break
                if not code_col:
                    code_col = code_columns[0]
            
            # Select best name column (prefer "final product names")
            if name_columns:
                for col in name_columns:
                    if 'final' in str(col).lower() and 'product' in str(col).lower():
                        name_col = col
                        break
                if not name_col:
                    for col in name_columns:
                        if 'product' in str(col).lower():
                            name_col = col
                            break
                if not name_col:
                    name_col = name_columns[0]

            if packaging_columns:
                packaging_col = packaging_columns[0]
            
            for index, row in df.iterrows():
                try:
                    # Get product code
                    if code_col and pd.notna(row[code_col]):
                        product_code = str(row[code_col]).strip()
                    else:
                        # Generate code from product name
                        if name_col and pd.notna(row[name_col]):
                            name = str(row[name_col]).strip()
                            # Create code from first 3 letters + row number
                            first_letters = ''.join([c for c in name if c.isalpha()][:3]).upper()
                            product_code = f"{first_letters}{index+1:03d}"
                        else:
                            product_code = f"PROD{index+1:03d}"
                    
                    # Get product name
                    if name_col and pd.notna(row[name_col]):
                        product_name = str(row[name_col]).strip()
                    else:
                        # Try to find any non-empty text column
                        for col in df.columns:
                            if col != code_col and pd.notna(row[col]):
                                val = str(row[col]).strip()
                                if len(val) > 3 and not val.isdigit():
                                    product_name = val
                                    break
                        else:
                            product_name = f"Product {index+1}"
                    
                    if product_code and product_name and product_code not in ['nan', 'Product Code']:
                        entry = {
                            'product_code': product_code,
                            'product_name': product_name,
                            'source_file': filename,
                        }
                        if packaging_col and pd.notna(row.get(packaging_col)):
                            entry['packaging_details'] = str(row[packaging_col]).strip()
                        products.append(entry)

                except Exception:
                    continue

    except Exception as e:
        hlog.warn("GENERATE", "manual extract failed", file=filename, reason=str(e))

    if products:
        hlog.info("GENERATE", "manual extract complete", file=filename, count=len(products))
    return products

def generate_certificate_with_products(
    certificate_no, company_name, company_reg_no, issue_date,
    certificate_type, standards, products, template_path, 
    output_png, output_pdf, text_color="black"
):
    """Generate certificate with extracted product information"""
    try:
        # Import the basic certificate generation function
        from certificate_generator import generate_certificate
        
        # Generate certificate with all information
        success = generate_certificate(
            certificate_no=certificate_no,
            company_name=company_name,
            company_reg_no=company_reg_no,
            issue_date=issue_date,
            template_path=template_path,
            output_png=output_png,
            output_pdf=output_pdf,
            text_color=text_color,
            products=products,
            certificate_type=certificate_type,
            standards=standards
        )
        
        if products:
            hlog.info("GENERATE", "certificate products bound", cert_no=certificate_no, count=len(products))

        return success

    except Exception as e:
        hlog.error("GENERATE", "certificate with products failed", reason=str(e))
        return False

def extract_certificate_number_regex_fallback(text):
    """Fallback regex-based certificate number extraction"""
    if not text:
        return None
    
    # Common patterns for certificate numbers (from most specific to most general)
    patterns = [
        # Most flexible patterns first - match certificate numbers with multiple slashes and varying lengths
        r'(?:download\s+certificate\s+|certificate\s*(?:no|number|#)?[:\s]*)([A-Z0-9/\-]{6,50})',  # HCO/RAO/TEST/202522323233
        r'(?:download\s+|get\s+|retrieve\s+)?(?:certificate\s+|cert\s+)?([A-Z]{2,4}/[A-Z0-9/\-]{4,40})',  # Very flexible with slashes
        # Original specific patterns
        r'(?:certificate\s*(?:no|number|#)?[:\s]*)?([A-Z]{2,4}/[A-Z]{2,4}/\d{6})',  # HCO/RAO/091024
        r'(?:certificate\s*(?:no|number|#)?[:\s]*)?([A-Z]{2,4}[-/]\d{4,6}[-/]\d{2,4})',  # HCO-2024-001
        r'(?:certificate\s*(?:no|number|#)?[:\s]*)?([A-Z]{2,4}\d{4,8})',  # HCO20240001
        r'(?:certificate\s*(?:no|number|#)?[:\s]*)?(\d{4,8}[-/]\d{2,4})',  # 2024-001
        r'(?:certificate\s*(?:no|number|#)?[:\s]*)?([A-Z]+\d+[A-Z]*\d*)',  # ABC123DEF456
        r'(?:certificate\s*(?:no|number|#)?[:\s]*)([A-Z0-9/\-]{6,30})',  # Generic alphanumeric with slashes/dashes
        r'(?:cert\s*(?:no|number|#)?[:\s]*)?([A-Z]{2,4}/[A-Z]{2,4}/\d{6})',  # cert HCO/RAO/091024
        r'(?:cert\s*(?:no|number|#)?[:\s]*)?([A-Z]{2,4}[-/]\d{4,6}[-/]\d{2,4})',  # cert HCO-2024-001
        r'(?:cert\s*(?:no|number|#)?[:\s]*)?([A-Z]{2,4}\d{4,8})',  # cert HCO20240001
        # More liberal patterns for simple certificate numbers
        r'(?:certificate\s*(?:no|number|#)?[:\s]*)?([a-zA-Z0-9/\-]{4,30})',  # Any alphanumeric with slashes/dashes
        r'(?:cert\s*(?:no|number|#)?[:\s]*)?([a-zA-Z0-9/\-]{4,30})',  # Any alphanumeric string with cert prefix
        # Ultra liberal pattern - extract any word that comes after certificate/cert (now includes slashes)
        r'(?:certificate|cert)\s+([a-zA-Z0-9/\-]+)',  # certificate iwhfoihwe or HCO/RAO/TEST/123
        r'download\s+certificate\s+([a-zA-Z0-9/\-]+)',  # download certificate HCO/RAO/TEST/123
        r'download\s+([a-zA-Z0-9/\-]+)',  # download HCO/RAO/TEST/123
    ]
    
    # Try each pattern
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        if matches:
            # Return the first match, cleaned up but preserve original case
            return matches[0].strip()
    
    return None

def analyze_certificate_query(query_text):
    """Analyze user query to determine intent: download, verification, or inquiry"""
    query_lower = query_text.lower()
    
    # Download keywords
    download_keywords = ['download', 'get certificate', 'retrieve certificate', 'get copy', 'download copy', 'file']
    
    # Verification keywords  
    verify_keywords = ['verify', 'check', 'validate', 'confirmation', 'authentic', 'valid', 'real']
    
    # Check for download intent
    for keyword in download_keywords:
        if keyword in query_lower:
            return "download"
    
    # Check for verification intent
    for keyword in verify_keywords:
        if keyword in query_lower:
            return "verification"
    
    # Check if certificate number is present without clear intent
    certificate_no = extract_certificate_number_from_text(query_text)
    if certificate_no:
        # If certificate number found but no clear intent, default to verification
        return "verification"
    
    # Default to inquiry for general questions
    return "inquiry"

def upload_certificate_to_supabase(certificate_no, file_data, file_type):
    """Upload certificate to Supabase storage and return public URL"""
    try:
        supabase_url = os.getenv("SUPABASE_URL")
        supabase_key = os.getenv("SUPABASE_ANON_KEY")
        
        if not supabase_url or not supabase_key:
            hlog.warn("APP", "supabase not configured")
            return None
        
        import requests
        
        bucket_name = "certificates"
        file_path = f"{certificate_no}.{file_type.lower()}"
        
        # Upload file to Supabase storage
        upload_url = f"{supabase_url}/storage/v1/object/{bucket_name}/{file_path}"
        
        headers = {
            "Authorization": f"Bearer {supabase_key}",
            "Content-Type": f"application/{file_type.lower()}"
        }
        
        response = requests.post(upload_url, data=file_data, headers=headers)
        
        if response.status_code in [200, 201]:
            # Return public URL
            public_url = f"{supabase_url}/storage/v1/object/public/{bucket_name}/{file_path}"
            return public_url
        else:
            hlog.warn("APP", "supabase upload failed", status=response.status_code)
            return None

    except Exception as e:
        hlog.warn("APP", "supabase upload error", reason=str(e))
        return None

def generate_supabase_download_url(certificate_no, file_type):
    """Generate secure download URL using Supabase"""
    try:
        # First, get the file data from database
        from database import get_certificate_file_from_db
        file_data = get_certificate_file_from_db(certificate_no, file_type)
        
        if not file_data:
            return None
        
        # Try to upload to Supabase and get public URL
        public_url = upload_certificate_to_supabase(certificate_no, file_data, file_type)
        
        if public_url:
            return public_url
        
        # Fallback: Generate direct database download URL through our API
        agent_url = os.getenv("AGENT_URL", "http://localhost:8025")
        fallback_url = f"{agent_url}/certificate/download"
        
        return f"{fallback_url}?certificate_no={certificate_no}&file_type={file_type}"
        
    except Exception as e:
        hlog.warn("APP", "download url generation failed", cert_no=certificate_no, reason=str(e))
        agent_url = os.getenv("AGENT_URL", "http://localhost:8025")
        return f"{agent_url}/certificate/download?certificate_no={certificate_no}&file_type={file_type}"

def verify_certificate(certificate_no):
    """
    Verify a certificate number without requiring end-user authentication.

    Source priority:
    Default:
      1) Database (or local fallback JSON via `database.get_certificate_from_db`)
      2) Excel (Microsoft Graph) table via `microsoft_graph` (optional)

    Override for testing:
      - Set HCO_VERIFY_EXCEL_FIRST=true to check Excel before local DB/fallback.
    """
    started_at = time.monotonic()
    try:
        excel_first = (os.getenv("HCO_VERIFY_EXCEL_FIRST") or "").strip().lower() in ("1", "true", "yes")
        hlog.verify_start(certificate_no, excel_first=excel_first)

        def _try_db() -> tuple[bool, dict]:
            try:
                from database import get_certificate_from_db
                db_cert = get_certificate_from_db(certificate_no)
                if db_cert:
                    hlog.verify_source(certificate_no, source="database", status="found")
                    return True, {
                        'certificate_no': db_cert.get('certificate_no', certificate_no),
                        'issue_date': db_cert.get('issue_date', ''),
                        'company_reg_no': db_cert.get('company_reg_no', ''),
                        'company_name': db_cert.get('company_name', ''),
                        'source': 'local_db_or_files',
                    }
                hlog.verify_source(certificate_no, source="database", status="not_found")
            except Exception as db_err:
                hlog.verify_source(certificate_no, source="database", status="error")
                hlog.warn("VERIFY", "database lookup failed", cert_no=certificate_no, reason=str(db_err))
            return False, {}

        def _try_excel() -> tuple[bool, dict]:
            try:
                ms_excel_share_url = os.getenv("HCO_EXCEL_SHARE_URL") or os.getenv("EXCEL_SHARE_URL")
                ms_table_name = (
                    os.getenv("HCO_EXCEL_TABLE_NAME")
                    or os.getenv("EXCEL_TABLE_NAME")
                    or "Certificates"
                )
                if not ms_excel_share_url:
                    hlog.verify_source(certificate_no, source="excel_graph", status="skipped")
                    return False, {}

                from microsoft_graph import get_access_token, find_row_in_excel_table_by_column_value

                token = get_access_token()
                row = find_row_in_excel_table_by_column_value(
                    ms_excel_share_url,
                    ms_table_name,
                    column_name="certificate_no",
                    match_value=certificate_no,
                    token=token,
                )
                if row:
                    hlog.verify_source(certificate_no, source="excel_graph", status="found")
                    return True, {
                        'certificate_no': row.get('certificate_no', certificate_no),
                        'issue_date': row.get('issue_date', ''),
                        'company_reg_no': row.get('company_reg_no', ''),
                        'company_name': row.get('company_name', ''),
                        'certificate_url': row.get('certificate_url', ''),
                        'source': 'excel_graph',
                    }
                hlog.verify_source(certificate_no, source="excel_graph", status="not_found")
            except Exception as excel_err:
                hlog.verify_source(certificate_no, source="excel_graph", status="error")
                hlog.warn("VERIFY", "excel lookup failed", cert_no=certificate_no, reason=str(excel_err))
            return False, {}

        attempts = (_try_excel, _try_db) if excel_first else (_try_db, _try_excel)
        for fn in attempts:
            ok, payload = fn()
            if ok:
                hlog.verify_done(
                    certificate_no,
                    verified=True,
                    duration_s=time.monotonic() - started_at,
                    source=payload.get("source", ""),
                    company=payload.get("company_name", ""),
                )
                return True, payload

        hlog.verify_done(certificate_no, verified=False, duration_s=time.monotonic() - started_at)
        return False, {}
    except Exception as e:
        hlog.error("VERIFY", "fatal", cert_no=certificate_no, reason=str(e))
        return False, {}

# Chat protocol message handler
@chat_proto.on_message(ChatMessage)
async def handle_message(ctx: Context, sender: str, msg: ChatMessage):
    ctx.logger.info(f"Got a message from {sender}")
    ctx.logger.info(f"Message content: {msg}")

    
    # Send acknowledgement
    await ctx.send(
        sender,
        ChatAcknowledgement(
            acknowledged_msg_id=msg.msg_id, 
            timestamp=datetime.now(timezone.utc)
        ),
    )

    prompt_content = []
    has_text_content = False
    has_valid_image_resource = False

    # First pass: collect content and identify what we have
    for item in msg.content:
        if isinstance(item, StartSessionContent):
            ctx.logger.info(f"Got a start session message from {sender}")
            # Signal that attachments are supported
            await ctx.send(sender, create_metadata({"attachments": "true"}))
        elif isinstance(item, TextContent):
            ctx.logger.info(f"Got text content from {sender}: {item.text}")
            has_text_content = True
            prompt_content.append({"text": item.text, "type": "text"})
        elif isinstance(item, ResourceContent):
            ctx.logger.info(f"Got resource content from {sender}")
            try:
                # Extract URI from resource content
                if item.resource:
                    # Handle both single Resource and list of Resources
                    resources = item.resource if isinstance(item.resource, list) else [item.resource]
                    
                    for resource in resources:
                        image_url = resource.uri
                        mime_type = resource.metadata.get('mime_type', 'image/jpeg')
                        # Fix generic 'image' MIME type to specific format
                        if mime_type == 'image':
                            mime_type = 'image/jpeg'
                        
                        ctx.logger.info(f"Extracted image URL: {image_url}")
                        ctx.logger.info(f"Mime type: {mime_type}")
                        
                        # Skip metadata resources that don't contain actual images
                        if image_url == 'metadata' or not image_url.startswith(('http', 'data:')):
                            ctx.logger.info(f"Skipping non-image resource: {image_url}")
                            continue
                        
                        # Only set flag if it's actually a valid image
                        has_valid_image_resource = True
                        prompt_content.append({
                            "type": "image_url",
                            "image_url": {"url": image_url},
                            "mime_type": mime_type,
                        })
                else:
                    ctx.logger.warning(f"No resource found in ResourceContent")
            except Exception as ex:
                ctx.logger.error(f"Failed to extract resource URI: {ex}")
                await ctx.send(sender, create_text_chat("Failed to extract image URL from resource."))
                return
        elif isinstance(item, MetadataContent):
            ctx.logger.info(f"Got metadata content from {sender}: {item.metadata}")
        else:
            ctx.logger.warning(f"Got unexpected content from {sender}: {type(item)}")

    # Log the final decision variables
    ctx.logger.info(f"Final decision: has_text_content={has_text_content}, has_valid_image_resource={has_valid_image_resource}")

    # Decision logic based on content types
    if has_text_content and has_valid_image_resource:
        # Both text and valid image resource present → always certificate verification
        ctx.logger.info("Both text and valid image resource present - processing as certificate verification")
        await process_certificate_verification(ctx, sender, prompt_content)
        return
    elif has_text_content and not has_valid_image_resource:
        # Only text content present → route through common text query processor
        text_messages = [item.get("text", "") for item in prompt_content if item.get("type") == "text"]
        full_text = " ".join(text_messages)

        try:
            qr = _process_text_query(full_text)
            ctx.logger.info(f"Chat protocol query processed as: {qr['query_type']}")
            await ctx.send(sender, create_text_chat(qr["message"]))
        except Exception as e:
            ctx.logger.error(f"Error processing text query in chat protocol: {e}")
            await ctx.send(sender, create_text_chat(f"Error processing your request: {e}"))
        return
    elif has_valid_image_resource and not has_text_content:
        # Only valid image resource present → certificate verification
        ctx.logger.info("Only valid image resource present - processing as certificate verification")
        await process_certificate_verification(ctx, sender, prompt_content)
        return
    else:
        # No meaningful content to process
        ctx.logger.info("No meaningful content to process")
        return

async def process_certificate_verification(ctx: Context, sender: str, prompt_content: list):
    """Process certificate verification for both text, image, and PDF content"""
    ctx.logger.info("Processing certificate verification content...")
    try:
        def _is_verifiable_resource(item: dict) -> bool:
            if item.get("type") == "image_url":
                return True
            if item.get("type") == "resource":
                mt = item.get("mime_type", "")
                return mt.startswith("image/") or mt.lower() in ("application/pdf", "pdf")
            return False

        has_verifiable = any(_is_verifiable_resource(item) for item in prompt_content)
        
        if has_verifiable:
            ctx.logger.info("Processing image/PDF for certificate verification...")
            # Process image for certificate extraction and validation
            response = get_image_analysis(prompt_content)
            await ctx.send(sender, create_text_chat(response))
            return
        else:
            # No image found – route through common text query processor
            text_messages = [item.get("text", "") for item in prompt_content if item.get("type") == "text"]
            full_text = " ".join(text_messages)
            ctx.logger.info(f"Processing text via common processor: {full_text}")
            qr = _process_text_query(full_text)
            await ctx.send(sender, create_text_chat(qr["message"]))
            return
            
    except Exception as err:
        ctx.logger.error(f"Error processing verification content: {err}")
        await ctx.send(sender, create_text_chat("Sorry, I couldn't process your request. Please try again later."))
        return

# Chat protocol acknowledgement handler
@chat_proto.on_message(ChatAcknowledgement)
async def handle_ack(ctx: Context, sender: str, msg: ChatAcknowledgement):
    ctx.logger.info(
        f"Got an acknowledgement from {sender} for {msg.acknowledged_msg_id}"
    )


@agent.on_rest_post("/image/upload", ImageRequest, ImageResponse)
async def handle_image_upload(ctx: Context, req: ImageRequest) -> ImageResponse:
    ctx.logger.info(f"Received request - Image: {req.filename is not None}, Text: {req.text_query is not None}")
    
    try:
        # Check if image data is provided
        if req.image_data and req.filename and req.content_type:
            # Handle image or PDF upload and analysis (always verification)
            ctx.logger.info(f"Processing file: {req.filename}, type: {req.content_type}")
            
            # Decode base64 image data
            image_bytes = base64.b64decode(req.image_data)
            image_size = len(image_bytes)
            
            ctx.logger.info(f"Image size: {image_size} bytes")
            
            # Extract certificate data using Haiku
            try:
                extracted_data = extract_certificate_data(req.image_data, req.content_type)
                ctx.logger.info(f"Certificate extraction result: {extracted_data}")
                
                # Check for extraction errors
                if "error" in extracted_data:
                    return ImageResponse(
                        message=f"Error extracting certificate data: {extracted_data['error']}",
                        agent_address=ctx.agent.address,
                        timestamp=int(time.time()),
                        image_size=image_size,
                        processed=False,
                    )
                
                # Validate certificate number without requiring user login:
                # - Prefer DB (or fallback JSON) and fall back to Google Sheets
                extracted_cert_no = extracted_data.get("certificate_no") if isinstance(extracted_data, dict) else None
                if not extracted_cert_no:
                    return ImageResponse(
                        message="Error extracting certificate number from image.",
                        agent_address=ctx.agent.address,
                        timestamp=int(time.time()),
                        image_size=image_size,
                        processed=False,
                    )

                is_valid, verified_data = verify_certificate(extracted_cert_no)
                ctx.logger.info(f"Certificate validation result (db/sheets): valid={is_valid}, data={verified_data}")
                
                # Format human-friendly response for image analysis
                if is_valid:
                    company_name = (verified_data.get('company_name') or extracted_data.get('company_name') or 'Unknown Company')
                    issue_date = (verified_data.get('issue_date') or extracted_data.get('issue_date') or 'Unknown Date')
                    certificate_no = verified_data.get('certificate_no') or extracted_cert_no
                    
                    response_message = f"✅ **Certificate Verified!**\n\n"
                    response_message += f"This is a valid HCO certificate for **{company_name}**, issued on **{issue_date}**.\n\n"
                    response_message += f"Certificate Number: {certificate_no}"
                else:
                    response_message = f"❌ **Certificate Not Valid**\n\n"
                    response_message += "This certificate is not valid. If you need a valid HCO certificate, please apply at https://www.hcoltd.co.uk/registration or contact HCO directly for assistance at info@hcoltd.co.uk or +44 (0) 333 577 0902."
                
                return ImageResponse(
                    message=response_message,
                    agent_address=ctx.agent.address,
                    timestamp=int(time.time()),
                    image_size=image_size,
                    processed=is_valid,
                )
                
            except Exception as validation_error:
                ctx.logger.error(f"Error during certificate validation: {str(validation_error)}")
                return ImageResponse(
                    message=f"Error validating certificate: {str(validation_error)}",
                    agent_address=ctx.agent.address,
                    timestamp=int(time.time()),
                    image_size=image_size,
                    processed=False,
                )
        
        elif req.text_query:
            # Handle text-only query with smart classification
            ctx.logger.info(f"Processing text query: {req.text_query}")
            
            # Classify the query first
            query_type = classify_query(req.text_query)
            ctx.logger.info(f"REST API Query classified as: {query_type}")
            
            if query_type == "inquiry":
                # Handle inquiry by searching HCO website first, then generating final answer
                ctx.logger.info("Processing inquiry query in REST API...")
                try:
                    search_results = search_hco_website(req.text_query)
                    final_answer = generate_final_answer(req.text_query, search_results)
                    
                    return ImageResponse(
                        message=final_answer,
                        agent_address=ctx.agent.address,
                        timestamp=int(time.time()),
                        image_size=0,
                        processed=True,
                    )
                except Exception as e:
                    ctx.logger.error(f"Error processing inquiry in REST API: {e}")
                    return ImageResponse(
                        message="I'm having trouble processing your inquiry right now. Please contact HCO directly at info@hcoltd.co.uk or +44 (0) 333 577 0902 for assistance.",
                        agent_address=ctx.agent.address,
                        timestamp=int(time.time()),
                        image_size=0,
                        processed=False,
                    )
            elif query_type == "download":
                # Handle download request
                ctx.logger.info("Processing download query in REST API...")
                # Extract certificate number from text
                certificate_no = extract_certificate_number_from_text(req.text_query)
                
                if certificate_no:
                    # Get certificate file from database
                    from database import get_certificate_file_from_db
                    file_data = get_certificate_file_from_db(certificate_no, "pdf")
                    
                    if file_data:
                        # Convert binary file data to base64
                        file_data_b64 = base64.b64encode(file_data).decode('utf-8')
                        
                        response = f"✅ **Certificate Downloaded Successfully!**\n\n"
                        response += f"📋 **Certificate Number:** {certificate_no}\n"
                        response += f"📄 **File:** {certificate_no}.pdf\n"
                        response += f"💾 **Status:** Downloaded to your device\n\n"
                        response += "Your certificate PDF has been downloaded and is ready to view."
                        
                        # Create a custom response that includes file data for download
                        return ImageResponse(
                            message=response,
                            agent_address=ctx.agent.address,
                            timestamp=int(time.time()),
                            image_size=0,
                            processed=True,
                            file_data=file_data_b64,
                            filename=f"{certificate_no}.pdf"
                        )
                    else:
                        response = f"❌ **Certificate Not Found**\n\n"
                        response += f"📋 **Certificate Number:** {certificate_no}\n\n"
                        response += "Certificate not found in our database. Please check the certificate number and try again."
                        
                        return ImageResponse(
                            message=response,
                            agent_address=ctx.agent.address,
                            timestamp=int(time.time()),
                            image_size=0,
                            processed=False,
                        )
                else:
                    response = "📋 **To download a certificate, please provide the certificate number.**\n\n"
                    response += "Examples:\n"
                    response += "• \"Download certificate HCO-2025-001\"\n"
                    response += "• \"Get certificate ABC-123\"\n"
                    response += "• \"I want to download certificate XYZ-456\"\n\n"
                    response += "I'll find and download your certificate PDF for you."
                    
                    return ImageResponse(
                        message=response,
                        agent_address=ctx.agent.address,
                        timestamp=int(time.time()),
                        image_size=0,
                        processed=False,
                    )
            elif query_type == "marketing":
                # Handle marketing content generation
                ctx.logger.info("Processing marketing content generation in REST API...")
                try:
                    marketing_content = generate_marketing_content(req.text_query)
                    
                    return ImageResponse(
                        message=marketing_content,
                        agent_address=ctx.agent.address,
                        timestamp=int(time.time()),
                        image_size=0,
                        processed=True,
                    )
                except Exception as e:
                    ctx.logger.error(f"Error generating marketing content in REST API: {e}")
                    return ImageResponse(
                        message="I'm having trouble generating marketing content right now. For marketing materials and promotional content, please contact HCO directly at info@hcoltd.co.uk or +44 (0) 333 577 0902.",
                        agent_address=ctx.agent.address,
                        timestamp=int(time.time()),
                        image_size=0,
                        processed=False,
                    )
            else:
                # Handle verification
                ctx.logger.info("Processing verification query in REST API...")
                # Extract certificate number from text
                certificate_no = extract_certificate_number_from_text(req.text_query)
                
                if certificate_no:
                    # Validate certificate against Google Sheets
                    is_valid, certificate_data = verify_certificate(certificate_no)
                    
                    if is_valid:
                        company_name = certificate_data.get('company_name', 'Unknown Company')
                        issue_date = certificate_data.get('issue_date', 'Unknown Date')
                        
                        response = f"✅ **Certificate Verified!**\n\n"
                        response += f"This is a valid HCO certificate for **{company_name}**, issued on **{issue_date}**.\n\n"
                        response += f"Certificate Number: {certificate_no}"
                        
                        return ImageResponse(
                            message=response,
                            agent_address=ctx.agent.address,
                            timestamp=int(time.time()),
                            image_size=0,
                            processed=True,
                        )
                    else:
                        response = f"❌ **Certificate Not Valid**\n\n"
                        response += "This certificate is not valid. If you need a valid HCO certificate, please apply at https://www.hcoltd.co.uk/registration or contact HCO directly for assistance at info@hcoltd.co.uk or +44 (0) 333 577 0902."
                        
                        return ImageResponse(
                            message=response,
                            agent_address=ctx.agent.address,
                            timestamp=int(time.time()),
                            image_size=0,
                            processed=False,
                        )
                else:
                    response = "I couldn't find a certificate number in your message. Please provide either:\n"
                    response += "• An image of the certificate to analyze\n"
                    response += "• A text message containing the certificate number (e.g., 'HCO-2024-001')"
                    
                    return ImageResponse(
                        message=response,
                        agent_address=ctx.agent.address,
                        timestamp=int(time.time()),
                        image_size=0,
                        processed=False,
                    )
        
        else:
            # Neither image nor text query provided
            return ImageResponse(
                message="Please provide either an image to analyze or a text query containing a certificate number.",
                agent_address=ctx.agent.address,
                timestamp=int(time.time()),
                image_size=0,
                processed=False,
            )
        
    except Exception as e:
        ctx.logger.error(f"Error processing request: {str(e)}")
        return ImageResponse(
            message=f"Error processing request: {str(e)}",
            agent_address=ctx.agent.address,
            timestamp=int(time.time()),
            image_size=0,
            processed=False,
        )

# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

# Aliases used to extract product code/name from heterogeneous product dicts
# (frontend payloads, docx parser output, sample data, etc.).
_PRODUCT_CODE_KEYS: tuple[str, ...] = (
    "product_code",
    "productcode",
    "code",
    "Product Code",
    "ProductCode",
    "sku",
    "item_code",
    "itemcode",
)
_PRODUCT_NAME_KEYS: tuple[str, ...] = (
    "description",
    "product_name",
    "productname",
    "Product Name",
    "ProductName",
    "Description",
    "product_description",
    "name",
)


def _pick_first_str(item: Dict[str, Any], keys: tuple[str, ...]) -> str:
    """Return the first non-empty stringified value found under *keys* (case-insensitive)."""
    if not isinstance(item, dict):
        return ""
    lower_map = {str(k).strip().lower(): v for k, v in item.items()}
    for key in keys:
        for candidate in (key, key.lower(), key.replace(" ", "_").lower()):
            val = lower_map.get(candidate)
            if val is None:
                continue
            text = str(val).strip()
            if text:
                return text
    return ""


def _build_products_csv_pair(products: Any) -> tuple[str, str]:
    """
    Build comma-separated `products_code` / `products_name` strings from a list of
    product dicts. Mirrors the aggregation done by the domestic generator so that
    the shared Excel/verification surface stays consistent.

    Edge cases:
    - Items missing one side (code or name) are still emitted on the side that is
      present; the other side is left blank for that index.
    - Duplicate values are deduplicated while preserving first-seen order.
    - Non-dict items are ignored.
    """
    if not isinstance(products, list):
        return "", ""

    codes: List[str] = []
    names: List[str] = []
    seen_codes: set[str] = set()
    seen_names: set[str] = set()

    for item in products:
        code = _pick_first_str(item, _PRODUCT_CODE_KEYS)
        name = _pick_first_str(item, _PRODUCT_NAME_KEYS)
        if not code and not name:
            continue
        if code and code not in seen_codes:
            codes.append(code)
            seen_codes.add(code)
        if name and name not in seen_names:
            names.append(name)
            seen_names.add(name)

    return ",".join(codes), ",".join(names)


def _excel_log_export(result: Dict[str, Any], req: "CertificateRequest", cert_type: str) -> None:
    """Best-effort append of export certificate metadata to the shared Excel table."""
    ms_excel_share_url = os.getenv("HCO_EXCEL_SHARE_URL") or os.getenv("EXCEL_SHARE_URL")
    ms_table_name = os.getenv("HCO_EXCEL_TABLE_NAME") or os.getenv("EXCEL_TABLE_NAME") or "Certificates"
    if not ms_excel_share_url:
        hlog.excel_skipped("HCO_EXCEL_SHARE_URL not configured")
        return
    try:
        from microsoft_graph import (
            get_access_token,
            get_excel_table_column_names,
            append_row_to_shared_excel_table,
        )
        token = get_access_token()
        uploaded_web_url = result.get("onedrive_web_url")
        row_data: Dict[str, Any] = {
            "certificate_id": result.get("certificate_id", ""),
            "certificate_no": req.certificate_no,
            "category": cert_type,
            "issue_date": req.issue_date,
            "company_reg_no": req.company_reg_no,
            "company_name": req.company_name,
            "certificate_url": uploaded_web_url,
            "created_at": datetime.now().isoformat(),
        }

        # Parity with domestic flow: populate product code/name aggregates so
        # downstream product verification (which reads from the Excel row) works
        # identically for export_non_meat. Generator result is preferred when
        # present so this also benefits export_meat if it ever returns products.
        products_code = str(result.get("products_code") or "").strip()
        products_name = str(result.get("products_name") or "").strip()
        if (not products_code and not products_name) and cert_type == "export_non_meat":
            products_code, products_name = _build_products_csv_pair(getattr(req, "products", None))

        if products_code or products_name:
            row_data["products_code"] = products_code
            row_data["products_name"] = products_name
            # Header alias support: some Excel tables use display-style headers.
            # Populating both alias keys keeps the schema unchanged while matching
            # whichever header style the workbook uses.
            row_data["Product Code"] = products_code
            row_data["Product Name"] = products_name

        col_names = get_excel_table_column_names(ms_excel_share_url, ms_table_name, token)
        if col_names:
            values = [row_data.get(name) for name in col_names]
        else:
            values = [
                row_data["certificate_id"],
                row_data["certificate_no"],
                row_data["issue_date"],
                row_data["company_reg_no"],
                row_data["company_name"],
                row_data.get("certificate_url") or row_data.get("created_at"),
            ]
        append_row_to_shared_excel_table(ms_excel_share_url, ms_table_name, values, token)
        hlog.excel_append(table=ms_table_name, cert_no=req.certificate_no, category=cert_type)
    except Exception as ms_err:
        hlog.warn("EXCEL", "append failed", table=ms_table_name, cert_no=req.certificate_no, reason=str(ms_err))


@agent.on_rest_post("/generate-certificate", CertificateRequest, CertificateResponse)
async def generate_certificate_endpoint(ctx: Context, req: CertificateRequest) -> CertificateResponse:
    ctx.logger.info(f"Received certificate generation request for {req.company_name}")
    
    try:
        # Restrict generation: require authentication (simple login or Microsoft token).
        # This keeps validation public but prevents anonymous generation.
        public_generation = (os.getenv("HCO_PUBLIC_GENERATION") or "").strip().lower() in ("1", "true", "yes")
        if not public_generation:
            token = (req.auth_token or "").strip()
            if not token:
                return CertificateResponse(
                    timestamp=int(time.time()),
                    message="❌ Generation denied: please log in.",
                    agent_address=ctx.agent.address,
                    certificate_id="",
                    png_filename="",
                    pdf_filename="",
                    download_url="",
                    csv_logged=False,
                    processed=False,
                )
            
            # Check if this is a simple login token
            if token.startswith("hco_token_"):
                ctx.logger.info("Simple login token detected for generation")
                try:
                    parts = token.split("_")
                    if len(parts) >= 4:
                        token_timestamp = int(parts[2])
                        username = "_".join(parts[3:])
                        
                        if int(time.time()) - token_timestamp > (24 * 3600):
                            return CertificateResponse(
                                timestamp=int(time.time()),
                                message="❌ Token expired. Please log in again.",
                                agent_address=ctx.agent.address,
                                certificate_id="",
                                png_filename="",
                                pdf_filename="",
                                download_url="",
                                csv_logged=False,
                                processed=False,
                            )
                        
                        ctx.logger.info(f"✅ Simple login token valid for {username}")
                    else:
                        raise ValueError("Invalid token format")
                except Exception as e:
                    ctx.logger.warning(f"Invalid simple login token: {str(e)}")
                    return CertificateResponse(
                        timestamp=int(time.time()),
                        message="❌ Invalid token. Please log in again.",
                        agent_address=ctx.agent.address,
                        certificate_id="",
                        png_filename="",
                        pdf_filename="",
                        download_url="",
                        csv_logged=False,
                        processed=False,
                    )
            else:
                # Microsoft token validation
                _check_user_can_access_onedrive_folder(token)
                email = _graph_me_email(token)
                allowed_emails = _parse_csv_env_list("HCO_ALLOWED_LOGIN_EMAILS")
                if allowed_emails and (not email or email.lower() not in allowed_emails):
                    return CertificateResponse(
                        timestamp=int(time.time()),
                        message="❌ Generation denied: your account is not allowed.",
                        agent_address=ctx.agent.address,
                        certificate_id="",
                        png_filename="",
                        pdf_filename="",
                        download_url="",
                        csv_logged=False,
                        processed=False,
                    )

        # HTML certificate generation doesn't need PNG template file

        # Check certificate type/category - handle export certificates differently
        cert_type = req.certificate_category or req.certificate_type or "halal_certificate"
        ctx.logger.info(f"Certificate type: {cert_type}")

        # Handle Export Certificates (Meat and Non-Meat)
        if cert_type in ("export_meat", "export_non_meat"):
            ctx.logger.info(f"Generating export certificate: {cert_type}")
            timings: Dict[str, float] = {}
            req_start = time.monotonic()

            export_data = {
                "certificate_no": req.certificate_no,
                "issue_date": req.issue_date,
                "standards": req.standards,
                "cert_num_footer": req.cert_num_footer,
                "country_of_origin": req.country_of_origin,
                "destination": req.destination,
                "exporter_name": req.exporter_name,
                "importer_name": req.importer_name,
                "export_logo_option": req.export_logo_option,
                "export_signature_option": req.export_signature_option,
            }

            if cert_type == "export_meat":
                export_data.update({
                    "slaughter_date": req.slaughter_date,
                    "expiry_date": req.expiry_date,
                    "abattoir_address": req.abattoir_address,
                    "gross_weight": req.gross_weight,
                    "number_of_carcasses": req.number_of_carcasses,
                    "net_weight": req.net_weight,
                    "number_of_boxes": req.number_of_boxes,
                    "batch_reference": req.batch_reference,
                    "halal_cert_number": req.halal_cert_number,
                    "vet_cert_number": req.vet_cert_number,
                    "destination_port": req.destination_port,
                    "loading_port": req.loading_port,
                    "flight_number": req.flight_number,
                    "meat_type": req.meat_type,
                    "awb_number": req.awb_number,
                    "meat_condition": req.meat_condition,
                    "inspector_name": req.inspector_name,
                })
            else:  # export_non_meat
                products = req.products
                if not isinstance(products, list):
                    products = []
                export_data.update({
                    "shipment_mode": req.shipment_mode,
                    "invoice_no": req.invoice_no,
                    "vet_health_cert_no": req.vet_health_cert_no,
                    "products": products,
                    "export_products_per_page": req.export_products_per_page,
                })

            # --- All export types (meat + non-meat): offload to background thread ---
            job_id = str(uuid4())
            _set_job(job_id, {
                "status": "queued",
                "message": "Certificate generation queued",
                "updated_at": time.time(),
            })
            ctx.logger.info(f"Queued {cert_type} generation job {job_id}")

            def _run_export_job(jid: str, etype: str, edata: dict, creq: CertificateRequest) -> None:
                jtimings: Dict[str, float] = {}
                job_started = time.monotonic()
                hlog.generate_start(job_id=jid, cert_no=creq.certificate_no, category=etype)
                _set_job(jid, {"status": "running", "message": "Generating PDF...", "updated_at": time.time()})
                try:
                    with step_timer("generate_pdf", jtimings):
                        result = generate_export_certificate(etype, edata)

                    if not result.get("success", False):
                        reason = result.get("error", "Unknown error")
                        hlog.generate_failed(job_id=jid, reason=reason, cert_no=creq.certificate_no)
                        _set_job(jid, {
                            "status": "failed",
                            "message": reason,
                            "updated_at": time.time(),
                            "timings": jtimings,
                        })
                        return

                    hlog.generate_step(
                        job_id=jid,
                        step="pdf_ready",
                        cert_id=result.get("certificate_id", ""),
                        onedrive=bool(result.get("onedrive_web_url")),
                    )

                    with step_timer("excel_log", jtimings):
                        _excel_log_export(result, creq, etype)

                    jtimings["total"] = sum(jtimings.values())
                    hlog.generate_done(
                        job_id=jid,
                        cert_no=creq.certificate_no,
                        duration_s=time.monotonic() - job_started,
                        cert_id=result.get("certificate_id", ""),
                    )
                    _set_job(jid, {
                        "status": "done",
                        "message": f"Export certificate generated for {creq.certificate_no}",
                        "certificate_id": result.get("certificate_id", ""),
                        "download_url": result.get("onedrive_web_url") or "",
                        "updated_at": time.time(),
                        "timings": jtimings,
                    })
                except Exception as exc:
                    hlog.generate_failed(job_id=jid, reason=str(exc), cert_no=creq.certificate_no)
                    _set_job(jid, {
                        "status": "failed",
                        "message": str(exc),
                        "updated_at": time.time(),
                        "timings": jtimings,
                    })

            thread = threading.Thread(
                target=_run_export_job,
                args=(job_id, cert_type, export_data, req),
                daemon=True,
            )
            thread.start()
            _gc_jobs()

            return CertificateResponse(
                timestamp=int(time.time()),
                message=f"{cert_type} certificate generation started. Poll /generation-status for progress.",
                agent_address=ctx.agent.address,
                certificate_id="",
                png_filename="",
                pdf_filename="",
                download_url="",
                csv_logged=False,
                processed=False,
                job_id=job_id,
                async_mode=True,
            )

        # Handle Slaughterhouse Certificate — async background thread
        if cert_type == "slaughterhouse":
            ctx.logger.info(f"Generating slaughterhouse certificate (async): {req.certificate_no}")

            from datetime import timedelta

            try:
                issue_dt = datetime.strptime(req.issue_date, '%Y-%m-%d')
            except Exception:
                issue_dt = datetime.now()

            validity_years = int(req.validity_period) if req.validity_period else 3
            expiry_date = (issue_dt + timedelta(days=365 * validity_years)).strftime('%Y-%m-%d')

            slaughter_params = dict(
                certificate_no=req.certificate_no,
                company_name=req.company_name,
                company_reg_no=req.company_reg_no,
                issue_date=req.issue_date,
                expiry_date=expiry_date,
                standards=req.standards,
                company_address=req.company_address,
                pu=req.pu,
                au=req.au,
                sow=req.sow,
                pl=req.pl,
                cert_num_footer=req.cert_num_footer,
                validity_period=req.validity_period,
                company_logo=req.company_logo,
            )

            job_id = str(uuid4())
            _set_job(job_id, {"status": "queued", "message": "Certificate generation queued", "updated_at": time.time()})
            ctx.logger.info(f"Queued slaughterhouse generation job {job_id}")

            def _run_slaughterhouse_job(jid: str, params: dict, creq: CertificateRequest) -> None:
                jtimings: Dict[str, float] = {}
                job_started = time.monotonic()
                hlog.generate_start(job_id=jid, cert_no=creq.certificate_no, category="slaughterhouse")
                _set_job(jid, {"status": "running", "message": "Generating PDF...", "updated_at": time.time()})
                try:
                    with step_timer("generate_pdf", jtimings):
                        result = generate_slaughterhouse_certificate(**params)

                    if not result.get("success", False):
                        reason = result.get("error", "Unknown error")
                        hlog.generate_failed(job_id=jid, reason=reason, cert_no=creq.certificate_no)
                        _set_job(jid, {
                            "status": "failed",
                            "message": reason,
                            "updated_at": time.time(),
                            "timings": jtimings,
                        })
                        return

                    hlog.generate_step(
                        job_id=jid,
                        step="pdf_ready",
                        cert_id=result.get("certificate_id", ""),
                        onedrive=bool(result.get("onedrive_web_url")),
                    )

                    with step_timer("excel_log", jtimings):
                        ms_excel_share_url = os.getenv("HCO_EXCEL_SHARE_URL") or os.getenv("EXCEL_SHARE_URL")
                        ms_table_name = os.getenv("HCO_EXCEL_TABLE_NAME") or os.getenv("EXCEL_TABLE_NAME") or "Certificates"
                        if not ms_excel_share_url:
                            hlog.excel_skipped("HCO_EXCEL_SHARE_URL not configured")
                        else:
                            try:
                                from microsoft_graph import get_access_token, get_excel_table_column_names, append_row_to_shared_excel_table
                                token = get_access_token()
                                row_data = {
                                    "certificate_id": result.get("certificate_id", ""),
                                    "certificate_no": creq.certificate_no,
                                    "category": "slaughterhouse",
                                    "issue_date": creq.issue_date,
                                    "company_reg_no": creq.company_reg_no,
                                    "company_name": creq.company_name,
                                    "certificate_url": result.get("onedrive_web_url"),
                                    "created_at": datetime.now().isoformat(),
                                }
                                col_names = get_excel_table_column_names(ms_excel_share_url, ms_table_name, token)
                                if col_names:
                                    values = [row_data.get(name) for name in col_names]
                                else:
                                    values = [row_data["certificate_id"], row_data["certificate_no"], row_data["issue_date"], row_data["company_reg_no"], row_data["company_name"], row_data.get("certificate_url") or row_data.get("created_at")]
                                append_row_to_shared_excel_table(ms_excel_share_url, ms_table_name, values, token)
                                hlog.excel_append(table=ms_table_name, cert_no=creq.certificate_no, category="slaughterhouse")
                            except Exception as ms_err:
                                hlog.warn("EXCEL", "append failed", table=ms_table_name, cert_no=creq.certificate_no, reason=str(ms_err))

                    jtimings["total"] = sum(jtimings.values())
                    hlog.generate_done(
                        job_id=jid,
                        cert_no=creq.certificate_no,
                        duration_s=time.monotonic() - job_started,
                        cert_id=result.get("certificate_id", ""),
                    )
                    _set_job(jid, {
                        "status": "done",
                        "message": f"Slaughterhouse certificate generated for {creq.company_name}",
                        "certificate_id": result.get("certificate_id", ""),
                        "download_url": result.get("onedrive_web_url") or "",
                        "updated_at": time.time(),
                        "timings": jtimings,
                    })
                except Exception as exc:
                    hlog.generate_failed(job_id=jid, reason=str(exc), cert_no=creq.certificate_no)
                    _set_job(jid, {"status": "failed", "message": str(exc), "updated_at": time.time(), "timings": jtimings})

            thread = threading.Thread(target=_run_slaughterhouse_job, args=(job_id, slaughter_params, req), daemon=True)
            thread.start()
            _gc_jobs()

            return CertificateResponse(
                timestamp=int(time.time()),
                message="Slaughterhouse certificate generation started. Poll /generation-status for progress.",
                agent_address=ctx.agent.address,
                certificate_id="",
                png_filename="",
                pdf_filename="",
                download_url="",
                csv_logged=False,
                processed=False,
                job_id=job_id,
                async_mode=True,
            )

        # Generate unique certificate ID for halal certificate
        certificate_id = str(uuid4())
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        ctx.logger.info(f"Generating halal certificate: {certificate_id}")

        # Generate certificate using new multi-page method
        # Using HTML certificate generation instead
        # from certificate_generator import generate_certificate_with_data
        from datetime import timedelta

        # Calculate expiry date based on validity period
        try:
            issue_dt = datetime.strptime(req.issue_date, '%Y-%m-%d')
        except Exception:
            issue_dt = datetime.now()

        # Calculate expiry date based on validity period (1, 2, or 3 years)
        validity_years = int(req.validity_period) if req.validity_period else 1
        expiry_date = (issue_dt + timedelta(days=365 * validity_years)).strftime('%Y-%m-%d')
        
        # Company logo functionality removed - no logo processing needed
        
        # Handle multiple XLSX/Excel files
        csv_files = None
        
        # First check for new xlsx_files format
        ctx.logger.info(f"🔍 XLSX files check: has xlsx_files={hasattr(req, 'xlsx_files')}, length={len(req.xlsx_files) if hasattr(req, 'xlsx_files') and req.xlsx_files else 0}")
        ctx.logger.info(f"🔍 XLSX files content: {req.xlsx_files}")
        ctx.logger.info(f"🔍 Request attributes: {[attr for attr in dir(req) if not attr.startswith('_')]}")
        ctx.logger.info(f"🔍 CSV files count: {req.csv_files_count}")
        
        # Log all non-empty fields to understand what the frontend is actually sending
        non_empty_fields = {}
        for attr in dir(req):
            if not attr.startswith('_') and hasattr(req, attr):
                value = getattr(req, attr)
                if value:  # Only log non-empty values
                    non_empty_fields[attr] = str(value)[:100]  # Truncate long values
        ctx.logger.info(f"🔍 Non-empty request fields: {non_empty_fields}")
        
        # Check if frontend is sending files via csv_files_count but not in the expected format
        if req.csv_files_count > 0:
            ctx.logger.info(f"🔍 Frontend indicates {req.csv_files_count} files but no data in expected fields")
            ctx.logger.info("🔍 This suggests a mismatch between frontend and backend file handling")
            
            # Look for any field that might contain file data
            for attr in dir(req):
                if not attr.startswith('_') and 'file' in attr.lower():
                    value = getattr(req, attr)
                    ctx.logger.info(f"🔍 File-related field '{attr}': {type(value)} = {str(value)[:100]}")
        
        if req.xlsx_files and len(req.xlsx_files) > 0:
            try:
                import base64
                csv_files = []
                
                # Create a mock file object that preserves binary data for Excel files
                class MockCSVFile:
                    def __init__(self, binary_content, filename):
                        self.filename = filename
                        self._binary_content = binary_content
                    
                    def read(self):
                        # Return binary content directly for Excel files
                        if self.filename.lower().endswith(('.xlsx', '.xls')):
                            return self._binary_content
                        else:
                            # For CSV files, try to decode with different encodings
                            for encoding in ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']:
                                try:
                                    return self._binary_content.decode(encoding)
                                except UnicodeDecodeError:
                                    continue
                            # Fallback for CSV
                            return self._binary_content.decode('latin-1')
                    
                    def seek(self, pos):
                        pass
                
                # Process each XLSX file
                for i, xlsx_file in enumerate(req.xlsx_files):
                    ctx.logger.info(f"🔍 Processing XLSX file {i}: {xlsx_file}")
                    if 'filename' in xlsx_file and 'data' in xlsx_file:
                        xlsx_data = base64.b64decode(xlsx_file['data'])
                        csv_files.append(MockCSVFile(xlsx_data, xlsx_file['filename']))
                        ctx.logger.info(f"✅ XLSX file processed: {xlsx_file['filename']}, size: {len(xlsx_data)} bytes")
                    else:
                        ctx.logger.warning(f"⚠️  Invalid XLSX file format: {xlsx_file}")
                
                ctx.logger.info(f"Total {len(csv_files)} XLSX files processed")
                
            except Exception as e:
                ctx.logger.error(f"Error processing XLSX files: {e}")
                # Don't fail the entire request due to file issues
                csv_files = None
        
        # Fallback to old single file format for backward compatibility
        ctx.logger.info(f"🔍 Legacy CSV check: has csv_file_data={bool(req.csv_file_data)}, has csv_file_filename={bool(req.csv_file_filename)}")
        if req.csv_file_data:
            ctx.logger.info(f"🔍 CSV file data length: {len(req.csv_file_data)}")
        if req.csv_file_filename:
            ctx.logger.info(f"🔍 CSV filename: {req.csv_file_filename}")
            
        if req.csv_file_data and req.csv_file_filename:
            try:
                import base64
                csv_data = base64.b64decode(req.csv_file_data)
                
                class MockCSVFile:
                    def __init__(self, binary_content, filename):
                        self.filename = filename
                        self._binary_content = binary_content
                    
                    def read(self):
                        if self.filename.lower().endswith(('.xlsx', '.xls')):
                            return self._binary_content
                        else:
                            for encoding in ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']:
                                try:
                                    return self._binary_content.decode(encoding)
                                except UnicodeDecodeError:
                                    continue
                            return self._binary_content.decode('latin-1')
                    
                    def seek(self, pos):
                        pass
                
                csv_files = [MockCSVFile(csv_data, req.csv_file_filename)]
                ctx.logger.info(f"Legacy single file processed: {req.csv_file_filename}")
            except Exception as e:
                ctx.logger.error(f"Error processing legacy file: {e}")
                csv_files = None
        
        # Format PU/AU text
        pu_au_text = ""
        if req.pu:
            pu_au_text = f"PU: {req.pu}"
        elif req.au:
            pu_au_text = f"AU: {req.au}"

        # --- Halal/Domestic: offload heavy work to background thread ---
        ctx.logger.info(f"🔍 Calling certificate generator with {len(csv_files) if csv_files else 0} CSV files")

        gen_params = dict(
            certificate_no=req.certificate_no,
            company_name=req.company_name,
            company_reg_no=req.company_reg_no,
            issue_date=req.issue_date,
            expiry_date=expiry_date,
            standards=req.standards,
            company_address=req.company_address,
            pu=req.pu,
            au=req.au,
            sow=req.sow,
            validity_period=req.validity_period,
            csv_files=csv_files,
            company_logo=req.company_logo,
            cert_num_footer=req.cert_num_footer,
            annex_layout_options=req.annex_layout_options,
            domestic_logo_1=req.domestic_logo_1,
            domestic_logo_2=req.domestic_logo_2,
        )

        job_id = str(uuid4())
        _set_job(job_id, {"status": "queued", "message": "Certificate generation queued", "updated_at": time.time()})
        ctx.logger.info(f"Queued {cert_type} generation job {job_id}")

        def _run_halal_domestic_job(jid: str, params: dict, cid: str, ctype: str, creq: CertificateRequest) -> None:
            jtimings: Dict[str, float] = {}
            job_started = time.monotonic()
            hlog.generate_start(job_id=jid, cert_no=creq.certificate_no, category=ctype)
            _set_job(jid, {"status": "running", "message": "Generating PDF...", "updated_at": time.time()})
            try:
                with step_timer("generate_pdf", jtimings):
                    result = generate_certificate_with_html_templates(**params)

                if not result.get("success", False):
                    reason = result.get("error", "Unknown error")
                    hlog.generate_failed(job_id=jid, reason=reason, cert_no=creq.certificate_no)
                    _set_job(jid, {
                        "status": "failed",
                        "message": reason,
                        "updated_at": time.time(),
                        "timings": jtimings,
                    })
                    return

                actual_certificate_id = result.get("certificate_id", cid)
                hlog.generate_step(
                    job_id=jid,
                    step="pdf_ready",
                    cert_id=actual_certificate_id,
                    onedrive=bool(result.get("onedrive_web_url")),
                )

                with step_timer("sheets_log", jtimings):
                    try:
                        save_to_sheets({
                            "certificate_id": actual_certificate_id,
                            "certificate_no": creq.certificate_no,
                            "company_name": creq.company_name,
                            "company_reg_no": creq.company_reg_no,
                            "issue_date": creq.issue_date,
                        }, "", "")
                    except Exception:
                        pass

                with step_timer("excel_log", jtimings):
                    ms_excel_share_url = os.getenv("HCO_EXCEL_SHARE_URL") or os.getenv("EXCEL_SHARE_URL")
                    ms_table_name = os.getenv("HCO_EXCEL_TABLE_NAME") or os.getenv("EXCEL_TABLE_NAME") or "Certificates"
                    if not ms_excel_share_url:
                        hlog.excel_skipped("HCO_EXCEL_SHARE_URL not configured")
                    else:
                        try:
                            from microsoft_graph import get_access_token, get_excel_table_column_names, append_row_to_shared_excel_table
                            token = get_access_token()
                            row_data = {
                                "certificate_id": actual_certificate_id,
                                "certificate_no": creq.certificate_no,
                                "category": ctype,
                                "issue_date": creq.issue_date,
                                "company_reg_no": creq.company_reg_no,
                                "company_name": creq.company_name,
                                "certificate_url": result.get("onedrive_web_url"),
                                "created_at": datetime.now().isoformat(),
                                "products_code": result.get("products_code") or "",
                                "products_name": result.get("products_name") or "",
                            }
                            col_names = get_excel_table_column_names(ms_excel_share_url, ms_table_name, token)
                            if col_names:
                                values = [row_data.get(name) for name in col_names]
                            else:
                                values = [row_data["certificate_id"], row_data["certificate_no"], row_data["issue_date"], row_data["company_reg_no"], row_data["company_name"], row_data.get("certificate_url") or row_data.get("created_at")]
                            append_row_to_shared_excel_table(ms_excel_share_url, ms_table_name, values, token)
                            hlog.excel_append(table=ms_table_name, cert_no=creq.certificate_no, category=ctype)
                        except Exception as ms_err:
                            hlog.warn("EXCEL", "append failed", table=ms_table_name, cert_no=creq.certificate_no, reason=str(ms_err))

                jtimings["total"] = sum(jtimings.values())
                hlog.generate_done(
                    job_id=jid,
                    cert_no=creq.certificate_no,
                    duration_s=time.monotonic() - job_started,
                    cert_id=actual_certificate_id,
                )
                _set_job(jid, {
                    "status": "done",
                    "message": f"Certificate generated for {creq.certificate_no}",
                    "certificate_id": actual_certificate_id,
                    "download_url": result.get("onedrive_web_url") or "",
                    "updated_at": time.time(),
                    "timings": jtimings,
                })
            except Exception as exc:
                hlog.generate_failed(job_id=jid, reason=str(exc), cert_no=creq.certificate_no)
                _set_job(jid, {"status": "failed", "message": str(exc), "updated_at": time.time(), "timings": jtimings})

        thread = threading.Thread(
            target=_run_halal_domestic_job,
            args=(job_id, gen_params, certificate_id, cert_type, req),
            daemon=True,
        )
        thread.start()
        _gc_jobs()

        return CertificateResponse(
            timestamp=int(time.time()),
            message=f"{cert_type} certificate generation started. Poll /generation-status for progress.",
            agent_address=ctx.agent.address,
            certificate_id="",
            png_filename="",
            pdf_filename="",
            download_url="",
            csv_logged=False,
            processed=False,
            job_id=job_id,
            async_mode=True,
        )
        
    except Exception as e:
        ctx.logger.error(f"Error generating certificate: {str(e)}")
        return CertificateResponse(
            timestamp=int(time.time()),
            message=f"Error generating certificate: {str(e)}",
            agent_address=ctx.agent.address,
            certificate_id="",
            png_filename="",
            pdf_filename="",
            download_url="",
            csv_logged=False,
            processed=False,
        )

@agent.on_rest_post("/generation-status", GenerationStatusRequest, GenerationStatusResponse)
async def generation_status_endpoint(ctx: Context, req: GenerationStatusRequest) -> GenerationStatusResponse:
    job = _get_job(req.job_id)
    if not job:
        return GenerationStatusResponse(
            timestamp=int(time.time()),
            job_id=req.job_id,
            status="not_found",
            message="No generation job found with this ID",
        )
    return GenerationStatusResponse(
        timestamp=int(time.time()),
        job_id=req.job_id,
        status=job.get("status", "unknown"),
        message=job.get("message", ""),
        certificate_id=job.get("certificate_id", ""),
        download_url=job.get("download_url", ""),
        timings=job.get("timings", {}),
    )


@agent.on_rest_post("/certificate/verify", CertificateVerifyRequest, CertificateVerifyResponse)
async def verify_certificate_endpoint(ctx: Context, req: CertificateVerifyRequest) -> CertificateVerifyResponse:
    ctx.logger.info(f"Received certificate verification request for {req.certificate_no}")
    
    try:
        is_valid, certificate_data = verify_certificate(req.certificate_no)
        
        if is_valid:
            company_name = certificate_data.get('company_name', 'Unknown Company')
            issue_date = certificate_data.get('issue_date', 'Unknown Date')
            
            message = f"✅ Certificate Verified! This is a valid HCO certificate for {company_name}, issued on {issue_date}. Certificate Number: {req.certificate_no}"
            
            ctx.logger.info(f"Certificate {req.certificate_no} is valid")
            return CertificateVerifyResponse(
                timestamp=int(time.time()),
                message=message,
                agent_address=ctx.agent.address,
                certificate_no=req.certificate_no,
                is_valid=True,
                certificate_data=certificate_data,
            )
        else:
            message = f"❌ Certificate Not Valid. This certificate is not valid. If you need a valid HCO certificate, please apply at https://www.hcoltd.co.uk/registration or contact HCO directly for assistance at info@hcoltd.co.uk or +44 (0) 333 577 0902."
            
            ctx.logger.info(f"Certificate {req.certificate_no} is invalid or not found")
            return CertificateVerifyResponse(
                timestamp=int(time.time()),
                message=message,
                agent_address=ctx.agent.address,
                certificate_no=req.certificate_no,
                is_valid=False,
                certificate_data={},
            )
        
    except Exception as e:
        ctx.logger.error(f"Error verifying certificate: {str(e)}")
        return CertificateVerifyResponse(
            timestamp=int(time.time()),
            message=f"Error verifying certificate: {str(e)}",
            agent_address=ctx.agent.address,
            certificate_no=req.certificate_no,
            is_valid=False,
            certificate_data={},
        )

@agent.on_rest_post("/certificate/verify-products", ProductVerifyRequest, ProductVerifyResponse)
async def verify_products_endpoint(ctx: Context, req: ProductVerifyRequest) -> ProductVerifyResponse:
    try:
        ok, payload = _verify_products_against_excel(
            certificate_no=req.certificate_no,
            product_names=req.product_names,
            product_codes=req.product_codes,
        )
        return ProductVerifyResponse(
            timestamp=int(time.time()),
            agent_address=ctx.agent.address,
            certificate_no=req.certificate_no,
            certificate_found=bool(payload.get("certificate_found")),
            verified=bool(payload.get("verified")),
            download_url=str(payload.get("download_url") or ""),
            message=str(payload.get("message") or ""),
            verified_product_names=list(payload.get("verified_product_names") or []),
            verified_product_codes=list(payload.get("verified_product_codes") or []),
            missing_product_names=list(payload.get("missing_product_names") or []),
            missing_product_codes=list(payload.get("missing_product_codes") or []),
        )
    except Exception as e:
        ctx.logger.error(f"Error verifying products: {str(e)}")
        return ProductVerifyResponse(
            timestamp=int(time.time()),
            message=f"Error verifying products: {str(e)}",
            agent_address=ctx.agent.address,
            certificate_no=req.certificate_no,
        )

@agent.on_rest_post("/chat", ChatRequest, ChatResponse)
async def handle_chat(ctx: Context, req: ChatRequest) -> ChatResponse:
    ctx.logger.info(f"Received chat request: {req.query}")
    try:
        # Classify the query first
        query_type = classify_query(req.query)
        ctx.logger.info(f"Chat REST API Query classified as: {query_type}")
        
        if query_type == "inquiry":
            # Handle inquiry by searching HCO website first, then generating final answer
            ctx.logger.info("Processing inquiry query in Chat REST API...")
            try:
                search_results = search_hco_website(req.query)
                final_answer = generate_final_answer(req.query, search_results)
                
                return ChatResponse(
                    timestamp=int(time.time()),
                    message=final_answer,
                    agent_address=ctx.agent.address,
                    query_type="inquiry",
                    processed=True,
                )
            except Exception as e:
                ctx.logger.error(f"Error processing inquiry in Chat REST API: {e}")
                return ChatResponse(
                    timestamp=int(time.time()),
                    message="I'm having trouble processing your inquiry right now. Please contact HCO directly at info@hcoltd.co.uk or +44 (0) 333 577 0902 for assistance.",
                    agent_address=ctx.agent.address,
                    query_type="inquiry",
                    processed=False,
                )
        elif query_type == "marketing":
            ctx.logger.info("Processing marketing content generation in Chat REST API...")
            try:
                marketing_content = generate_marketing_content(req.query)
                return ChatResponse(
                    timestamp=int(time.time()),
                    message=marketing_content,
                    agent_address=ctx.agent.address,
                    query_type="marketing",
                    processed=True,
                )
            except Exception as e:
                ctx.logger.error(f"Error generating marketing content in Chat REST API: {e}")
                return ChatResponse(
                    timestamp=int(time.time()),
                    message="I'm having trouble generating marketing content right now. For marketing materials and promotional content, please contact HCO directly at info@hcoltd.co.uk or +44 (0) 333 577 0902.",
                    agent_address=ctx.agent.address,
                    query_type="marketing",
                    processed=False,
                )
        else:
            # Handle verification
            ctx.logger.info("Processing verification query in Chat REST API...")
            # Extract certificate number from text
            certificate_no = extract_certificate_number_from_text(req.query)
            
            if certificate_no:
                # Validate certificate against Google Sheets
                is_valid, certificate_data = verify_certificate(certificate_no)
                
                if is_valid:
                    company_name = certificate_data.get('company_name', 'Unknown Company')
                    issue_date = certificate_data.get('issue_date', 'Unknown Date')
                    
                    response = f"✅ **Certificate Verified!**\n\n"
                    response += f"This is a valid HCO certificate for **{company_name}**, issued on **{issue_date}**.\n\n"
                    response += f"Certificate Number: {certificate_no}"
                    
                    return ChatResponse(
                        timestamp=int(time.time()),
                        message=response,
                        agent_address=ctx.agent.address,
                        query_type="verification",
                        processed=True,
                    )
                else:
                    response = f"❌ **Certificate Not Valid**\n\n"
                    response += "This certificate is not valid. If you need a valid HCO certificate, please apply at https://www.hcoltd.co.uk/registration or contact HCO directly for assistance at info@hcoltd.co.uk or +44 (0) 333 577 0902."
                    
                    return ChatResponse(
                        timestamp=int(time.time()),
                        message=response,
                        agent_address=ctx.agent.address,
                        query_type="verification",
                        processed=False,
                    )
            else:
                response = "I couldn't find a certificate number in your message. Please provide either:\n"
                response += "• An image of the certificate to analyze\n"
                response += "• A text message containing the certificate number (e.g., 'HCO-2024-001')"
                
                return ChatResponse(
                    timestamp=int(time.time()),
                    message=response,
                    agent_address=ctx.agent.address,
                    query_type="verification",
                    processed=False,
                )
        
    except Exception as e:
        ctx.logger.error(f"Error processing chat request: {str(e)}")
        return ChatResponse(
            timestamp=int(time.time()),
            message=f"Error processing request: {str(e)}",
            agent_address=ctx.agent.address,
            query_type="unknown",
            processed=False,
            certificate_no=None,
            certificate_found=False,
            verified_product_names=[],
            verified_product_codes=[],
            missing_product_names=[],
            missing_product_codes=[],
        )

@agent.on_rest_post("/certificate/query", CertificateQueryRequest, CertificateQueryResponse)
async def handle_certificate_query(ctx: Context, req: CertificateQueryRequest) -> CertificateQueryResponse:
    """Handle certificate queries with analysis for download, verification, or inquiry"""
    ctx.logger.info(f"Received certificate query: {req.query}")
    try:
        # Analyze the query to determine intent
        query_type = classify_query(req.query)
        ctx.logger.info(f"Query analyzed as: {query_type}")
        
        # Extract certificate number if present
        certificate_no = extract_certificate_number_from_text(req.query)
        
        if query_type == "download":
            if certificate_no:
                ctx.logger.info(f"Processing download request for certificate: {certificate_no}")
                
                # Check if certificate exists in database
                from database import get_certificate_from_db, get_certificate_file_from_db
                certificate_data = get_certificate_from_db(certificate_no)
                
                if certificate_data:
                    # Try to get PDF first, fallback to PNG
                    file_data = get_certificate_file_from_db(certificate_no, "pdf")
                    file_type = "pdf"
                    
                    if not file_data:
                        file_data = get_certificate_file_from_db(certificate_no, "png")
                        file_type = "png"
                    
                    if file_data:
                        # Return success message - frontend will handle download via /download-certificate endpoint
                        filename = f"{certificate_no}.{file_type}"
                        download_url = f"http://localhost:8025/download-certificate"
                        
                        return CertificateQueryResponse(
                            timestamp=int(time.time()),
                            message=f"✅ **Certificate Found!**\n\n📋 **Certificate Number:** {certificate_no}\n📄 **File:** {filename}\n💾 **Status:** Ready for download",
                            agent_address=ctx.agent.address,
                            query_type="download",
                            certificate_no=certificate_no,
                            download_url=download_url,
                            processed=True,
                            filename=filename,
                            found=True
                        )
                    else:
                        return CertificateQueryResponse(
                            timestamp=int(time.time()),
                            message=f"Certificate {certificate_no} found but no file data available for download.",
                            agent_address=ctx.agent.address,
                            query_type="download",
                            certificate_no=certificate_no,
                            download_url=None,
                            processed=False,
                            filename=None,
                            found=False
                        )
                else:
                    return CertificateQueryResponse(
                        timestamp=int(time.time()),
                        message=f"Certificate {certificate_no} not found in database.",
                        agent_address=ctx.agent.address,
                        query_type="download",
                        certificate_no=certificate_no,
                        download_url=None,
                        processed=False,
                        filename=None,
                        found=False
                    )
            else:
                return CertificateQueryResponse(
                    timestamp=int(time.time()),
                    message="Please provide a certificate number to download.",
                    agent_address=ctx.agent.address,
                    query_type="download",
                    certificate_no=None,
                    download_url=None,
                    processed=False,
                    filename=None,
                    found=False
                )
        
        elif query_type == "verification":
            if certificate_no:
                ctx.logger.info(f"Processing verification request for certificate: {certificate_no}")
                
                # Verify certificate
                is_valid, certificate_data = verify_certificate(certificate_no)
                
                if is_valid:
                    company_name = certificate_data.get('company_name', 'Unknown Company')
                    issue_date = certificate_data.get('issue_date', 'Unknown Date')
                    
                    message = f"✅ **Certificate Verified!**\n\n"
                    message += f"This is a valid HCO certificate for **{company_name}**, issued on **{issue_date}**.\n\n"
                    message += f"Certificate Number: {certificate_no}"
                    
                    return CertificateQueryResponse(
                        timestamp=int(time.time()),
                        message=message,
                        agent_address=ctx.agent.address,
                        query_type="verification",
                        certificate_no=certificate_no,
                        download_url=None,
                        processed=True,
                        filename=None,
                        found=True
                    )
                else:
                    message = f"❌ **Certificate Not Valid**\n\n"
                    message += "This certificate is not valid. If you need a valid HCO certificate, please apply at https://www.hcoltd.co.uk/registration or contact HCO directly for assistance at info@hcoltd.co.uk or +44 (0) 333 577 0902."
                    
                    return CertificateQueryResponse(
                        timestamp=int(time.time()),
                        message=message,
                        agent_address=ctx.agent.address,
                        query_type="verification",
                        certificate_no=certificate_no,
                        download_url=None,
                        processed=False,
                        filename=None,
                        found=False
                    )
            else:
                return CertificateQueryResponse(
                    timestamp=int(time.time()),
                    message="Please provide a certificate number to verify.",
                    agent_address=ctx.agent.address,
                    query_type="verification",
                    certificate_no=None,
                    download_url=None,
                    processed=False,
                    filename=None,
                    found=False
                )
        
        elif query_type == "marketing":
            ctx.logger.info(f"Processing marketing content generation: {req.query}")
            
            try:
                marketing_content = generate_marketing_content(req.query)
                
                return CertificateQueryResponse(
                    timestamp=int(time.time()),
                    message=marketing_content,
                    agent_address=ctx.agent.address,
                    query_type="marketing",
                    certificate_no=None,
                    download_url=None,
                    processed=True,
                    filename=None,
                    found=False
                )
            except Exception as e:
                ctx.logger.error(f"Error generating marketing content: {e}")
                return CertificateQueryResponse(
                    timestamp=int(time.time()),
                    message="I'm having trouble generating marketing content right now. For marketing materials and promotional content, please contact HCO directly at info@hcoltd.co.uk or +44 (0) 333 577 0902.",
                    agent_address=ctx.agent.address,
                    query_type="marketing",
                    certificate_no=None,
                    download_url=None,
                    processed=False,
                    filename=None,
                    found=False
                )
        
        elif query_type == "product_verification":
            ctx.logger.info(f"Processing product verification: {req.query}")
            
            if certificate_no:
                try:
                    payload = _do_product_verification(certificate_no, req.query)
                    return CertificateQueryResponse(
                        timestamp=int(time.time()),
                        message=str(payload.get("message") or ""),
                        agent_address=ctx.agent.address,
                        query_type="product_verification",
                        certificate_no=certificate_no,
                        download_url=str(payload.get("download_url") or "") or None,
                        processed=bool(payload.get("verified")),
                        filename=None,
                        found=bool(payload.get("certificate_found")),
                        verified_product_names=list(payload.get("verified_product_names") or []),
                        verified_product_codes=list(payload.get("verified_product_codes") or []),
                        missing_product_names=list(payload.get("missing_product_names") or []),
                        missing_product_codes=list(payload.get("missing_product_codes") or []),
                    )
                except Exception as e:
                    ctx.logger.error(f"Error verifying products: {e}")
                    return CertificateQueryResponse(
                        timestamp=int(time.time()),
                        message=f"Error verifying products for certificate {certificate_no}. Please try again or contact HCO for assistance.",
                        agent_address=ctx.agent.address,
                        query_type="product_verification",
                        certificate_no=certificate_no,
                        download_url=None,
                        processed=False,
                        filename=None,
                        found=False
                    )
            else:
                return CertificateQueryResponse(
                    timestamp=int(time.time()),
                    message="Please provide a certificate number along with the product names or codes to verify.\n\n**Example:** _verify product names Beef-XP, Chicken 500g for certificate HCO-2024-001_",
                    agent_address=ctx.agent.address,
                    query_type="product_verification",
                    certificate_no=None,
                    download_url=None,
                    processed=False,
                    filename=None,
                    found=False
                )
        
        elif query_type == "generation":
            ctx.logger.info(f"Processing certificate generation request: {req.query}")
            
            # Generation requests should redirect to certificate generation UI
            return CertificateQueryResponse(
                timestamp=int(time.time()),
                message="🔒 **Certificate Generation Available**\n\nTo generate a new certificate, please use the step-by-step certificate generation interface. Simply type **'generate certificate'** in the chat to start the guided process.\n\nFor assistance with certificate generation, contact HCO at info@hcoltd.co.uk or +44 (0) 333 577 0902.",
                agent_address=ctx.agent.address,
                query_type="generation",
                certificate_no=None,
                download_url=None,
                processed=True,
                filename=None,
                found=False
            )
        
        else:  # inquiry
            ctx.logger.info(f"Processing inquiry: {req.query}")
            
            try:
                search_results = search_hco_website(req.query)
                final_answer = generate_final_answer(req.query, search_results)
                
                return CertificateQueryResponse(
                    timestamp=int(time.time()),
                    message=final_answer,
                    agent_address=ctx.agent.address,
                    query_type="inquiry",
                    certificate_no=None,
                    download_url=None,
                    processed=True,
                    filename=None,
                    found=False
                )
            except Exception as e:
                ctx.logger.error(f"Error processing inquiry: {e}")
                return CertificateQueryResponse(
                    timestamp=int(time.time()),
                    message="I'm having trouble processing your inquiry right now. Please contact HCO directly at info@hcoltd.co.uk or +44 (0) 333 577 0902 for assistance.",
                    agent_address=ctx.agent.address,
                    query_type="inquiry",
                    certificate_no=None,
                    download_url=None,
                    processed=False,
                    filename=None,
                    found=False
                )
        
    except Exception as e:
        ctx.logger.error(f"Error processing certificate query: {str(e)}")
        return CertificateQueryResponse(
            timestamp=int(time.time()),
            message=f"Error processing query: {str(e)}",
            agent_address=ctx.agent.address,
            query_type="unknown",
            certificate_no=None,
            download_url=None,
            processed=False,
            filename=None,
            found=False,
        )

@agent.on_rest_post("/certificate/download", CertificateDownloadRequest, CertificateDownloadResponse)
async def download_certificate(ctx: Context, req: CertificateDownloadRequest) -> CertificateDownloadResponse:
    ctx.logger.info(f"Received certificate download request for {req.certificate_no}, type: {req.file_type}")
    
    # Ensure certificate_no is not None
    cert_no = (req.certificate_no or "").strip()
    if not cert_no:
        ctx.logger.error("Missing certificate_no in download request")
        return CertificateDownloadResponse(
            certificate_no="",
            file_data=None,
            filename=None,
            download_url=None,
            found=False,
            message="Missing certificate number"
        )
    
    try:
        # First, try to get the download URL from Excel (most reliable)
        excel_share_url = os.getenv("HCO_EXCEL_SHARE_URL") or os.getenv("EXCEL_SHARE_URL")
        excel_table_name = (
            os.getenv("HCO_EXCEL_TABLE_NAME")
            or os.getenv("EXCEL_TABLE_NAME")
            or "Certificates"
        )
        folder_share_url = (os.getenv("HCO_ONEDRIVE_FOLDER_SHARE_URL") or os.getenv("ONEDRIVE_FOLDER_SHARE_URL") or "").strip()
        
        if excel_share_url and excel_table_name:
            try:
                from microsoft_graph import get_access_token, find_row_in_excel_table_by_column_value
                token = get_access_token()
                
                # Find the certificate row in Excel
                cert_row = find_row_in_excel_table_by_column_value(
                    excel_share_url,
                    excel_table_name,
                    column_name="certificate_no",
                    match_value=cert_no,
                    token=token,
                )
                
                if cert_row:
                    # Certificate found in Excel - fetch a fresh download URL (stored URL may have expired token)
                    safe_cert_no = cert_no.replace("/", "_").replace("\\", "_")
                    filename = f"certificate_{safe_cert_no}.{req.file_type.lower()}"
                    
                    # Get a fresh download URL from OneDrive (don't use stored URL as it may have expired tempauth)
                    if folder_share_url:
                        try:
                            from microsoft_graph import get_shared_folder_file_web_url
                            fresh_download_url = get_shared_folder_file_web_url(folder_share_url, filename, token)
                            ctx.logger.info(f"✅ Found certificate in Excel, fetched fresh download URL for {cert_no}")
                            response = CertificateDownloadResponse(
                                certificate_no=cert_no,
                                file_data=None,
                                filename=filename,
                                download_url=fresh_download_url,
                                found=True,
                                message="Certificate found; returning fresh OneDrive download link.",
                            )
                            ctx.logger.info(f"✅ Returning download response: found={response.found}, download_url={response.download_url[:50] if response.download_url else 'None'}...")
                            return response
                        except Exception as e:
                            ctx.logger.warning(f"Could not fetch fresh download URL, trying stored URL: {e}")
                            # Fallback to stored URL if fresh fetch fails
                            stored_url = cert_row.get("certificate_url") or ""
                            if stored_url:
                                ctx.logger.info(f"⚠️  Using stored URL (may be expired): {stored_url[:50]}...")
                                return CertificateDownloadResponse(
                                    certificate_no=cert_no,
                                    file_data=None,
                                    filename=filename,
                                    download_url=stored_url,
                                    found=True,
                                    message="Certificate found; returning stored OneDrive link (may require authentication).",
                                )
            except Exception as e:
                ctx.logger.warning(f"Excel lookup failed, trying OneDrive folder: {e}")
        
        # Fallback: try to construct URL from OneDrive folder
        safe_cert_no = cert_no.replace("/", "_").replace("\\", "_")
        filename = f"certificate_{safe_cert_no}.{req.file_type.lower()}"

        if folder_share_url:
            try:
                from microsoft_graph import get_access_token, get_shared_folder_file_web_url
                token = get_access_token()
                web_url = get_shared_folder_file_web_url(folder_share_url, filename, token)
                return CertificateDownloadResponse(
                    certificate_no=cert_no,
                    file_data=None,
                    filename=filename,
                    download_url=web_url,
                    found=True,
                    message="Returning OneDrive link for certificate file",
                )
            except Exception as e:
                ctx.logger.warning(f"OneDrive download link lookup failed, falling back to DB/local file: {e}")

        from database import get_certificate_file_from_db
        
        # Get file data from database
        file_data = get_certificate_file_from_db(cert_no, req.file_type)
        
        if file_data:
            # Generate a safe filename
            # (keep consistent with OneDrive upload naming)
            # safe_cert_no/filename already computed above.

            # Avoid returning huge base64 blobs (uAgents schema/transport can fail on big payloads).
            max_inline_bytes = int(os.getenv("MAX_INLINE_DOWNLOAD_BYTES", "120000"))

            if len(file_data) > max_inline_bytes:
                # If OneDrive folder is configured, return a web link instead of inline base64
                try:
                    if folder_share_url:
                        from microsoft_graph import get_access_token, get_shared_folder_file_web_url
                        token = get_access_token()
                        web_url = get_shared_folder_file_web_url(folder_share_url, filename, token)
                        return CertificateDownloadResponse(
                            certificate_no=cert_no,
                            file_data=None,
                            filename=filename,
                            download_url=web_url,
                            found=True,
                            message="Certificate file is large; returning OneDrive download link",
                        )
                except Exception as e:
                    ctx.logger.warning(f"Large download fallback to OneDrive link failed: {e}")
                # If we can't produce a link, fall through and try inline base64 anyway.

            # Encode file data as base64
            file_data_b64 = base64.b64encode(file_data).decode("utf-8")

            return CertificateDownloadResponse(
                certificate_no=cert_no,
                file_data=file_data_b64,
                filename=filename,
                download_url=None,
                found=True,
                message=f"Certificate {req.file_type.upper()} file found and ready for download",
            )
        else:
            return CertificateDownloadResponse(
                certificate_no=cert_no,
                file_data=None,
                filename=None,
                download_url=None,
                found=False,
                message=f"Certificate {req.file_type.upper()} file not found for {cert_no}"
            )
        
    except Exception as e:
        ctx.logger.error(f"Error downloading certificate: {str(e)}", exc_info=True)
        return CertificateDownloadResponse(
            certificate_no=cert_no,
            file_data=None,
            filename=None,
            download_url=None,
            found=False,
            message=f"Error downloading certificate: {str(e)}"
        )


@agent.on_rest_post("/export-non-meat/parse-docx", ParseNonMeatDocxRequest, ParseNonMeatDocxResponse)
async def parse_export_non_meat_docx(ctx: Context, req: ParseNonMeatDocxRequest) -> ParseNonMeatDocxResponse:
    try:
        public_generation = (os.getenv("HCO_PUBLIC_GENERATION") or "").strip().lower() in ("1", "true", "yes")
        if not public_generation:
            token = (req.auth_token or "").strip()
            if not token:
                return ParseNonMeatDocxResponse(
                    timestamp=int(time.time()),
                    message="❌ Parsing denied: please log in.",
                    agent_address=ctx.agent.address,
                    products=[],
                    processed=False,
                )

            if token.startswith("hco_token_"):
                pass
            else:
                _check_user_can_access_onedrive_folder(token)

        if not (req.file_data or "").strip():
            return ParseNonMeatDocxResponse(
                timestamp=int(time.time()),
                message="Missing DOCX file_data.",
                agent_address=ctx.agent.address,
                products=[],
                processed=False,
            )

        try:
            from docx import Document
        except Exception as e:
            return ParseNonMeatDocxResponse(
                timestamp=int(time.time()),
                message=f"DOCX parsing dependency missing: {e}",
                agent_address=ctx.agent.address,
                products=[],
                processed=False,
            )

        raw = base64.b64decode(req.file_data)
        doc = Document(BytesIO(raw))

        def _norm(s: str) -> str:
            return re.sub(r"[^a-z0-9]+", "", (s or "").strip().lower())

        def _parse_date_to_iso(value: str) -> str:
            v = (value or "").strip()
            if not v:
                return ""
            v = v.replace("\n", " ").strip()
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%d/%m/%y", "%d-%m-%y"):
                try:
                    return datetime.strptime(v, fmt).date().isoformat()
                except Exception:
                    pass

            # Handle Month Year formats (e.g., "September 2025"), defaulting day=01
            for fmt in ("%B %Y", "%b %Y", "%B-%Y", "%b-%Y", "%B/%Y", "%b/%Y"):
                try:
                    dt = datetime.strptime(v, fmt)
                    return dt.replace(day=1).date().isoformat()
                except Exception:
                    pass
            m = re.search(r"(\d{1,2})[\-/.](\d{1,2})[\-/.](\d{2,4})", v)
            if m:
                d, mth, y = m.group(1), m.group(2), m.group(3)
                if len(y) == 2:
                    y = "20" + y
                try:
                    return datetime.strptime(f"{int(d):02d}/{int(mth):02d}/{y}", "%d/%m/%Y").date().isoformat()
                except Exception:
                    return ""
            return ""

        def _split_two_dates(value: str) -> tuple[str, str]:
            v = (value or "").strip()
            if not v:
                return "", ""
            parts = [p.strip() for p in re.split(r"\s*/\s*", v) if p.strip()]
            if len(parts) >= 2:
                return _parse_date_to_iso(parts[0]), _parse_date_to_iso(parts[1])
            return _parse_date_to_iso(v), ""

        def _split_two_weights(value: str) -> tuple[str, str]:
            v = (value or "").strip()
            if not v:
                return "", ""
            parts = [p.strip() for p in re.split(r"\s*/\s*", v) if p.strip()]
            if len(parts) >= 2:
                return parts[0].replace("kg", "").strip(), parts[1].replace("kg", "").strip()
            return v.replace("kg", "").strip(), ""

        header_keys = {
            "productcode": {"productcode", "code", "itemcode", "sku"},
            "description": {
                "description",
                "productname",
                "productdescription",
                "name",
                "descriptionofshippedproduct",
                "descriptionofshippedgoods",
                "descriptionofgoods",
            },
            "quantity": {"quantity", "qty", "quantityunits"},
            "manufacture_date": {"manufacturedate", "mfgdate", "mfg"},
            "expiry_date": {"expirydate", "expdate", "exp"},
            "mfgexp": {
                "mfgexp",
                "mfgexpiry",
                "manufactureexpiry",
                "manufactureexpirydate",
                "manufacturedateexpirydate",
                "manufacturedateexpirydates",
                "date",
                "dates",
            },
            "batch_number": {"batchnumber", "batch", "lot"},
            "gross_weight": {"grossweight", "gross"},
            "net_weight": {"netweight", "net"},
            "grossnet": {"grossnet", "weight", "weights", "totalgrossweightnetweight"},
            "number_of_cases": {
                "numberofcases",
                "cases",
                "noofcases",
                "noofcase",
                "cartons",
                "casesnumberofpacks",
            },
        }

        def _find_idx(norm_headers: List[str], keys: set[str]) -> int:
            for i, h in enumerate(norm_headers):
                if h in keys:
                    return i
            return -1

        extracted: List[Dict[str, Any]] = []

        for table in getattr(doc, "tables", []) or []:
            rows: List[List[str]] = []
            for r in table.rows:
                row_cells = [c.text.strip() for c in r.cells]
                if any((cell or "").strip() for cell in row_cells):
                    rows.append(row_cells)
            if not rows:
                continue

            header_row_idx = -1
            col_map: Dict[str, int] = {}

            for i, row in enumerate(rows[:25]):
                norm_headers = [_norm(c) for c in row]
                desc_idx = _find_idx(norm_headers, header_keys["description"])
                if desc_idx == -1:
                    continue
                header_row_idx = i
                col_map["description"] = desc_idx
                col_map["product_code"] = _find_idx(norm_headers, header_keys["productcode"])
                col_map["quantity"] = _find_idx(norm_headers, header_keys["quantity"])
                col_map["manufacture_date"] = _find_idx(norm_headers, header_keys["manufacture_date"])
                col_map["expiry_date"] = _find_idx(norm_headers, header_keys["expiry_date"])
                col_map["mfgexp"] = _find_idx(norm_headers, header_keys["mfgexp"])
                col_map["batch_number"] = _find_idx(norm_headers, header_keys["batch_number"])
                col_map["gross_weight"] = _find_idx(norm_headers, header_keys["gross_weight"])
                col_map["net_weight"] = _find_idx(norm_headers, header_keys["net_weight"])
                col_map["grossnet"] = _find_idx(norm_headers, header_keys["grossnet"])
                col_map["number_of_cases"] = _find_idx(norm_headers, header_keys["number_of_cases"])
                break

            if header_row_idx == -1:
                continue

            for row in rows[header_row_idx + 1 :]:
                def _get(idx: int) -> str:
                    if idx is None or idx < 0:
                        return ""
                    if idx >= len(row):
                        return ""
                    return (row[idx] or "").strip()

                description = _get(col_map.get("description", -1))
                if not description:
                    continue

                mfg = ""
                exp = ""
                if col_map.get("mfgexp", -1) >= 0:
                    mfg, exp = _split_two_dates(_get(col_map.get("mfgexp", -1)))
                else:
                    mfg = _parse_date_to_iso(_get(col_map.get("manufacture_date", -1)))
                    exp = _parse_date_to_iso(_get(col_map.get("expiry_date", -1)))

                gross = ""
                net = ""
                if col_map.get("grossnet", -1) >= 0:
                    gross, net = _split_two_weights(_get(col_map.get("grossnet", -1)))
                else:
                    gross = _get(col_map.get("gross_weight", -1)).replace("kg", "").strip()
                    net = _get(col_map.get("net_weight", -1)).replace("kg", "").strip()

                extracted.append(
                    {
                        "product_code": _get(col_map.get("product_code", -1)),
                        "description": description,
                        "quantity": _get(col_map.get("quantity", -1)),
                        "manufacture_date": mfg,
                        "expiry_date": exp,
                        "batch_number": _get(col_map.get("batch_number", -1)),
                        "gross_weight": gross,
                        "net_weight": net,
                        "number_of_cases": _get(col_map.get("number_of_cases", -1)),
                    }
                )

        if not extracted:
            return ParseNonMeatDocxResponse(
                timestamp=int(time.time()),
                message="No product table detected in the uploaded DOCX.",
                agent_address=ctx.agent.address,
                products=[],
                processed=False,
            )

        return ParseNonMeatDocxResponse(
            timestamp=int(time.time()),
            message=f"Extracted {len(extracted)} product row(s) from DOCX.",
            agent_address=ctx.agent.address,
            products=extracted,
            processed=True,
        )

    except Exception as e:
        ctx.logger.error(f"Error parsing non-meat DOCX: {str(e)}", exc_info=True)
        return ParseNonMeatDocxResponse(
            timestamp=int(time.time()),
            message=f"Error parsing DOCX: {str(e)}",
            agent_address=ctx.agent.address,
            products=[],
            processed=False,
        )

# Register protocols
agent.include(chat_proto, publish_manifest=True)

# CORS is handled natively by uagents framework

if __name__ == "__main__":
    hlog.configure()
    hlog.info("APP", "boot start")
    initialize_csv()

    try:
        init_database()
        hlog.info("APP", "database ready")
    except Exception as e:
        hlog.warn("APP", "database init failed", reason=str(e))

    hlog.info("APP", "agent listening", port=os.getenv("AGENT_PORT", "8096"))
    agent.run()


    